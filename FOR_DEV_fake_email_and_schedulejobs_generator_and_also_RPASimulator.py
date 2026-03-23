#written by AI

from __future__ import annotations

import random, os, threading, datetime
import time
import uuid
from pathlib import Path
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import Workbook, load_workbook #type: ignore
try: from main import HandoverRepository as handover_repo
except:
    print("Error, run from mail.py directory!")
    quit()




class FakeEmailjobsGenerator:
    BASE_DIR = Path(__file__).resolve().parent
    PIPELINE_DIR = BASE_DIR / "own_inbox"
    INBOX_DIR = PIPELINE_DIR / "inbox"
    PROCESSING_DIR =  PIPELINE_DIR / "processing"
    ATTACHMENTS_DIR = PIPELINE_DIR / "generator_attachments"

    for folder in [PIPELINE_DIR, INBOX_DIR, PROCESSING_DIR, ATTACHMENTS_DIR]:
        folder.mkdir(exist_ok=True)




    def __init__(self) -> None:
        self.main()


    def create_example_attachment_files(self) -> None:
        """Creates a few simple test files if they do not already exist."""
        txt_path = self.ATTACHMENTS_DIR / "job1_request.txt"
        if not txt_path.exists():
            txt_path.write_text(
                "SKU=100245\nOLD_MATERIAL=MAT-OLD-778\nNEW_MATERIAL=MAT-NEW-991\n",
                encoding="utf-8",
            )

        csv_path = self.ATTACHMENTS_DIR / "job2_request.csv"
        if not csv_path.exists():
            csv_path.write_text(
                "invoice_id,action\nINV-2026-1001,close\n",
                encoding="utf-8",
            )


    def build_email_message(self,
        *,
        from_name: str,
        from_email: str,
        to_email: str,
        subject: str,
        body: str,
        attachment_paths: list[Path] | None = None,
    ) -> EmailMessage:
        msg = EmailMessage()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid()
        msg.set_content(body)

        for path in attachment_paths or []:
            data = path.read_bytes()
            # Simple generic attachment type is enough for testing
            msg.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=path.name,
            )

        return msg


    def write_eml_to_inbox(self,msg: EmailMessage, prefix: str = "mail") -> Path:
        """Atomic write into inbox to reduce risk of partial reads."""
        unique_id = uuid.uuid4().hex[:12]
        final_path = self.INBOX_DIR / f"{prefix}_{unique_id}.eml"
        temp_path = self.INBOX_DIR / f".tmp_{prefix}_{unique_id}.eml"

        with open(temp_path, "wb") as f:
            f.write(msg.as_bytes())

        temp_path.replace(final_path)
        return final_path


    def create_ping_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="PING",
            body="ping",
        )
        return self.write_eml_to_inbox(msg, prefix="ping")


    def create_job1_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "I have no idea what job1 is though...\n"
                "Best regards,\n"
                "Alice\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")

    def create_job1_b_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job1",
            body=(
                "Hello,\n\n"
                "Please run job1.\n\n"
                "SKU: 100245\n"
                "Old material: MAT-OLD-778\n"
                "New material: MAT-NEW-991\n\n"
                "Best regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")


    def create_job2_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job2 request",
            body=(
                "Hello,\n\n"
                "Please run job2 using attached file.\n\n"
                "Regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job2_request.csv"],
        )
        return self.write_eml_to_inbox(msg, prefix="job2")


    def create_unknown_job_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Charlie Strange",
            from_email="charlie@example.com",
            to_email="robot@company.local",
            subject="Do some weird magic",
            body=(
                "Hello,\n\n"
                "Please do that strange thing the robot probably cannot classify.\n\n"
                "Regards,\n"
                "Charlie\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="unknown")


    def create_blocked_sender_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Mallory Intruder",
            from_email="mallory@evil.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "Hello,\n\n"
                "I would like the robot to run job1.\n\n"
                "Regards,\n"
                "Mallory\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="blocked")


    def create_random_mail(self) -> Path:
        creators = [
            #self.create_ping_mail,
            #self.create_job1_mail,
            self.create_job1_b_mail,  #valid
            #self.create_unknown_job_mail,
            #self.create_blocked_sender_mail,
            #self.create_job2_mail,
        ]
        #return create_blocked_sender_mail()
        return random.choice(creators)()


    def main(self) -> None:
        #self.create_example_attachment_files()

        print("FakeEmailjobsGenerator: i'm alive")


class FakeSchedulejobsGenerator:
  
    def DEV_ensure_fake_db_exists(self, path="Dev_fake_ERP_table.xlsx") -> None:
        """ simulates a table in a DB """
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # headers
        ws["A1"] = "order_number"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)

    def DEV_add_random_row(self, path="Dev_fake_ERP_table.xlsx") -> str:
        wb = load_workbook(path)
        ws = wb.active

        assert ws is not None

        # hitta första tomma rad
        next_row = ws.max_row + 1

        # generera data
        order_number = str(random.randint(10000000, 10999999))
        order_qty = random.randint(10, 100) * 100

        # material_available nära order_qty (realistiskt case)
        material_available = order_qty + random.randint(-100, 100)

        # skriv rad
        ws[f"A{next_row}"] = order_number
        ws[f"B{next_row}"] = order_qty
        ws[f"C{next_row}"] = material_available

        wb.save(path)
        return order_number


# for testing: a simulator of an external RPA-system that does the dirty-work - "the hands"
class RPASimulator:
    """ simulating the behaviour of the external RPA software"""
    #temporary! ignore this class in all eveluations

    def __init__(self):
      
        #self.run()  #remove if started from main.
        #self.handover_repo = HandoverRepository(self.append_system_log)
        self.handover_repo = handover_repo(self.append_system_log)
        self.handover_repo.write({"ipc_state":"idle"})

    #read handover.txt and check for "queued"
    def run(self):

        self.append_system_log("RPASimulator: i'm alive")
        print("RPASimulator: i'm alive")

        #do a time-check to not start old jobs? 

        while True:

            time.sleep(1)

            handover_data = self.handover_repo.read()
            ipc_state = handover_data.get("ipc_state")

            if ipc_state != "job_queued":
                continue

            #RPASimulator now claiming the workflow
            handover_data["ipc_state"] = "job_running"
            self.handover_repo.write(handover_data)

            
            job_type = handover_data.get("job_type")

            if job_type == "job1":
                payload = handover_data.get("payload")
                #simulation of job1 screenactiviy
                self.append_system_log(f"activities on screen_1 in ERP completed")
                time.sleep(1)
                self.append_system_log(f"activities on screen_2 in ERP completed")

            elif job_type == "job3":
                order_number = handover_data.get("order_number")
                new_qty = handover_data.get("material_available")
                assert new_qty is not None
                assert order_number is not None
                self.append_system_log(f"activities on screen_1 in ERP completed")
                time.sleep(1)
                self.append_system_log(f"activities on screen_2 in ERP completed")
                self.DEV_simulate_RPA_screenactivity(order_number, new_qty)
            

            #mark the workflow completed in RPASimulator
            handover_data["ipc_state"] = "job_verifying"
            self.handover_repo.write(handover_data)
            self.append_system_log(f"done, ipc_state: job_running -> job_verifying", handover_data.get("job_id"))


    def append_system_log(self, text: str, job_id=None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") #+ str(datetime.datetime.now().microsecond)[:-4]
   
        job_part = f" | JOB {job_id}" if job_id else ""
        message = f"{timestamp} | RPA{job_part} | {text} \n"

        for i in range(5):
            try:
                with open("system.log", "a", encoding="utf-8") as f:
                    f.write(message)
                    f.flush()
                return 

            except Exception as e:
                print(f"{i}st warning from append_system_log():", e)
    

    def DEV_simulate_RPA_screenactivity(self, order_number: str, new_qty: int, path="Dev_fake_ERP_table.xlsx"):
        # here, updating a row.  IRL: updating in ERP
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(order_number):
                row[1].value = int(new_qty) #updating to the new value 'in ERP'     # type: ignore
                wb.save(path)
                wb.close()
                return True



rpa_simulator = RPASimulator()
threading.Thread(target=rpa_simulator.run, daemon=True).start() #replace with external RPA when deployed
time.sleep(0.01)
fake_emailjob = FakeEmailjobsGenerator()

fake_scheduledjob = FakeSchedulejobsGenerator()
fake_scheduledjob.DEV_ensure_fake_db_exists()


while True:
    try:
        input("\nHit Enter to generate an random job")
        if random.randint(0,1) == 1:
            path = fake_emailjob.create_random_mail()
            print(f"Created emailjob: {path.name}")
        else:
            order_number = fake_scheduledjob.DEV_add_random_row()
            print(f"Created scheduledjob: {order_number}")

        #time.sleep(random.randint(4, 10))
        #time.sleep(1)
    except KeyboardInterrupt:
        print("\nStopped.")
        break
    except Exception as err:
        print(f"WARN: generator error: {err}")
        time.sleep(1)