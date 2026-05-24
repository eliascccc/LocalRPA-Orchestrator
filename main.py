import atexit
import datetime
import json
import os
import platform
import re
import shutil
import signal
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import tkinter as tk
from dataclasses import asdict, dataclass
from email import policy
from email.parser import BytesParser
from email.utils import parseaddr
from pathlib import Path
from typing import Any, Literal, Never, TypeAlias, get_args
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore

VERSION = "0.4"
CONFIG_FILE = "robotruntime_config.json"


# ============================================================
# DATA MODELS
# ============================================================

HandoverState: TypeAlias = Literal["idle", "job_queued", "job_running", "job_verifying", "safestop"]
JobName: TypeAlias = str    # or JobName: TypeAlias = Literal["ping", "qty_adjust", "po_adjust", "order_adjust", ... ]
SourceType: TypeAlias = Literal["personal_inbox", "shared_inbox", "erp_query"]
DashboardStatus: TypeAlias = Literal["online", "safestop", "working", "no_network" , "out_of_office"]
LifecycleStatus: TypeAlias = Literal["REJECTED", "QUEUED", "RUNNING", "VERIFYING", "FAIL", "DONE"]
LifecycleErrorCode: TypeAlias = Literal["PRE_HANDOVER_CRASH", "RPA_TOOL_CRASH", "POST_HANDOVER_VERIFICATION_MISMATCH",  "POST_HANDOVER_VERIFICATION_TIMEOUT", "POST_HANDOVER_UNSPEC_CRASH", "OUT_OF_SERVICE", "OUTSIDE_WORKING_HOURS", "UNKNOWN_JOB", "NO_ACCESS", "NO_NETWORK", "INVALID_INPUT", "CODE_ERROR", "RECOVERY_SOURCE_MISSING", "IN_SAFESTOP"]
RuntimePhase: TypeAlias = Literal["startup", "poll_intake", "personal_precheck", "shared_precheck", "query_precheck", "queue_for_rpa", "waiting_rpa_claim", "waiting_rpa_execution", "verification",]

@dataclass
class HandoverFile:
    """Data stored in handover.json and exchanged with the RPA tool."""
    
    state: HandoverState                            # eg. "job_queued"
    job_name: JobName | None = None                 # eg. "qty_adjust"
    job_id: int | None = None                       # eg. 202611051223
    rpatool_payload: dict[str, Any] | None = None   # eg. {"order_number": 12345, "target_qty": 44, "pick_qty_from_location": "WH7",} (this is the final data sent to RPA Tool)

@dataclass
class ActiveJob:
    '''History, recovery, reply, verification context'''

    source_ref: str                                 # backend identifier eg. Outlook EntryID or "ERP_ORDER:12345" (demo uses filename for email backend)
    source_type: SourceType                         # eg. "personal_inbox"
    source_data: dict[str, Any] | None = None       # eg. {"order_number": 12345, "requested_qty": 44}
    rpatool_payload: dict[str, Any] | None = None   # eg. {"order_number": 12345, "target_qty": 44, "location": "WH7",}
    request_summary: str | None = None              # eg. "Change PO 12345 from 43 pcs to 44 pcs"
    job_name: JobName | None = None                 # eg. "qty_adjust"

    job_id: int | None = None               # runtime identifier eg. 202611481132

    email_address: str | None = None        # for email
    email_subject: str | None = None        # for email
    email_body: str | None = None           # for email eg. "Hi, change the order 12345 to 44 pcs"

@dataclass
class PrecheckResult:
    is_success: bool
    error_code: LifecycleErrorCode | None = None
    public_error_message: str | None = None
    request_summary: str | None = None
    rpatool_payload: dict[str, Any] | None = None

@dataclass
class VerificationResult:
    is_success: bool
    error_code: LifecycleErrorCode | None = None
    public_error_message: str | None = None

class RobotRuntimeFault(Exception):
    def __init__(self, phase: RuntimePhase, message:str|None=None,  handover_file:HandoverFile|None=None, active_job: ActiveJob | None = None, error_code: LifecycleErrorCode = "CODE_ERROR", cause:Exception|None=None, traceback_text:str|None=None):
        super().__init__(message)
        self.error_message = message
        self.phase: RuntimePhase = phase
        self.active_job = active_job
        self.handover_file = handover_file
        self.error_code: LifecycleErrorCode = error_code
        self.cause = cause
        self.traceback_text = traceback_text


@dataclass(frozen=True)
class RuntimeConfig:
    # Default demo values. After first run, edit robotruntime_config.json with local values.

    rpa_tool_claim_timeout: int = 10        # Max seconds RR waits for RPA tool to pick up a queued job
    rpa_tool_execution_timeout: int = 10    # (demo friendly) max seconds for RPA tool to finish workflow
    poll_interval: int = 1                  # (demo-friendly) poll interval for runtime_loop()
    queryflow_poll_interval: int = 1        # (demo-friendly) poll interval for query_flow

    operating_hours_start: int = 5 # eg. 05:00
    operating_hours_end: int = 23
    operating_days: tuple[int, ...] = (0, 1, 2, 3, 4, 5, 6) # Monday=0, Tuesday=1, ..., Saturday=5, Sunday=6

    system_log_path: str = "system.log"
    handover_file: str = "handover.json"
    audit_db_path: str = "job_audit.db"
    friends_path: str = "friends.xlsx"

    recordings_in_progress_folder: str = "recordings_in_progress"
    recordings_destination_folder: str = "recordings_destination" # A (demo-friendly) local destination, move to a shared drive accessabile to all users
    network_healthcheck_path: str | None = None    # None is a (demo-friendly) always healthy path, replace with eg "G:\\"
    
    erp_system_name: str = "M3"
    organisation_domain: str = "@example.com"
    rpa_admin_email: str = "ada.lovelace@example.com"
    public_robot_email: str = "robot@example.com"
    mail_command_job_id: int = 999999999999

def load_or_create_config(path: str = CONFIG_FILE):
    # written by AI
    default_config = RuntimeConfig()

    if not os.path.exists(path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(asdict(default_config), f, indent=2)
        return default_config

    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    allowed_keys = set(RuntimeConfig.__dataclass_fields__.keys())

    unknown_keys = set(raw.keys()) - allowed_keys
    if unknown_keys:
        raise ValueError(f"Unknown config keys in {path}: {sorted(unknown_keys)}")

    return RuntimeConfig(**raw)


if __name__ == "__main__": sys.modules.setdefault("main", sys.modules[__name__]) # avoid double import of dataclasses

# ============================================================
# BACKENDS
# ============================================================

class DemoMailBackend:
    """Demo mailbox simulated with local folders and .eml files (replace w/ eg. Outlook)"""

    MAIL_STATUS_PREFIXES = ("PROCESSING", "DONE", "FAIL")

    def __init__(self, source_type) -> None:
        self.source_type = source_type
        self.inbox_dir = Path(self.source_type) / "inbox"
        self.inbox_dir.mkdir(parents=True, exist_ok=True)

    def list_inbox_mail_paths(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.inbox_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str
        return paths

    def parse_mail_file(self, mail_path) -> ActiveJob:
        with open(mail_path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        from_name, from_address = parseaddr(msg.get("From", ""))
        del from_name # not used

        email_address = (from_address or "").strip().lower()
        if not email_address or "@" not in email_address:
            email_address = None

        email_subject = msg.get("Subject", "").strip()

        # not needed in demo
        # message_id = msg.get("Message-ID", "").strip()    # eg. Outlook EntryID / Graph ID in real backend
        # raw_headers = {k: str(v) for k, v in msg.items()} # good for troubleshooting metadata 

        if msg.is_multipart():
            body_parts = []
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass
            email_body = "\n".join(body_parts).strip()
        else:
            try:
                email_body = msg.get_content().strip()
            except Exception:
                email_body = ""
        

        attachments = {}
        # placeholder for implementation

        return ActiveJob(
            source_ref=mail_path,
            email_address=email_address,
            email_subject=email_subject,
            email_body=email_body,
            source_type=self.source_type,
            source_data=attachments,
            )

    def mark_processing(self, active_job: ActiveJob) -> ActiveJob:
        original_subject = self._strip_status_prefix(active_job.email_subject)
        new_subject = f"PROCESSING/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(active_job, new_subject)

    def mark_done(self, active_job: ActiveJob) -> ActiveJob:
        original_subject = self._strip_status_prefix(active_job.email_subject)
        new_subject = f"DONE/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(active_job, new_subject)

    def mark_failed(self, active_job: ActiveJob) -> ActiveJob:
        original_subject = self._strip_status_prefix(active_job.email_subject)
        new_subject = f"FAIL/{self._today_yyyymmdd()}/{original_subject}"
        return self._set_subject(active_job, new_subject)

    def send_reply(self, active_job: ActiveJob, extra_subject: str, extra_body: str) -> None:

        reply_to = active_job.email_address

        original_subject = self._strip_status_prefix(active_job.email_subject)
        subject = f"{extra_subject} re: {original_subject}"

        body = (
            f"{extra_body} \n\n"
            f"-------------------------------------------------------------\n"
            f"{active_job.email_body}"
        ) # In a real mail backend, this should use the native reply mechanism.
        
        if reply_to is None:
            raise ValueError("cannot send reply because active_job.email_address is None")
        
        self._print_email_preview(reply_to, subject, body)
                    
    def delete(self, active_job: ActiveJob, fallback_status: Literal["DONE", "FAIL"]) -> None:
        try:
            os.remove(active_job.source_ref)

        except Exception as err:
            if (self._has_status_prefix(active_job, "DONE") or self._has_status_prefix(active_job, "FAIL")):
                # delete failed, but mail already has final status prefix
                return

            if fallback_status == "DONE":
                self.mark_done(active_job)

            elif fallback_status == "FAIL":
                self.mark_failed(active_job)

    def sent_reply_exists(self, source_ref) -> bool:
        '''an extra check in real backend sent folder, used to avoid duble user reply in error handling'''
        
        # placeholder for implementaton
        # if active_job.source_ref exists in personal sent mail from robot, return True
        # (job_id is present in all sent mails, except the 'online notice' mail, as 'tag' under robot_signature)
        return False

    def _today_yyyymmdd(self) -> str:
        return datetime.datetime.now().strftime("%Y%m%d")

    def _strip_status_prefix(self, subject: str | None) -> str:
        subject = (subject or "").strip()

        for status in self.MAIL_STATUS_PREFIXES:
            pattern = rf"^{status}/\d{{8}}/(.*)$"
            match = re.match(pattern, subject, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip()

        return subject

    def _has_status_prefix(self, active_job: ActiveJob, status: str | None = None) -> bool:
        subject = (active_job.email_subject or "").strip()

        if status is not None:
            return bool(re.match(rf"^{status}/\d{{8}}/", subject, flags=re.IGNORECASE))

        return bool(re.match(r"^(PROCESSING|DONE|FAIL)/\d{8}/", subject, flags=re.IGNORECASE))

    def _set_subject(self, active_job: ActiveJob, new_subject: str) -> ActiveJob:
        """ Demo backend: update the .eml file, real Outlook backend: use native subject rename."""
        path = Path(active_job.source_ref)

        with open(path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        if "Subject" in msg:
            msg.replace_header("Subject", new_subject)
        else:
            msg["Subject"] = new_subject

        with open(path, "wb") as f:
            f.write(msg.as_bytes(policy=policy.default))

        active_job.email_subject = new_subject

        return active_job

    def _print_email_preview(self, reply_to: str, subject: str, body: str):

        print(
        "\n" + "="*72 +
        "\n📧 EMAIL REPLY PREVIEW\n" +
        "="*72 +
        f"\nFrom:    robot@backend_example.com"
        f"\nTo:      {reply_to}"
        f"\nSubject: {subject}"
        f"\nDate:    {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        "\n" + "-"*72 +
        f"\n{body}\n" +
        "="*72 + "\n"
    )


class DemoErpBackend:
    """Demo ERP backend simulated with an Excel file."""

    def order_adjust_selection_rows(self, path="Demo_ERP_table.xlsx") -> list[dict]:
        '''this demo query is triggered by a job in custom_queryjobs.py'''


        self._ensure_demo_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active

        assert ws is not None #to satisfy pylance

        all_rows=[]

        for row in ws.iter_rows(min_row=2):  # skip header
            
            source_ref = row[0].value
            order_qty = row[1].value
            material_available = row[2].value

            if order_qty != material_available:

                all_rows.append({
                        "source_ref": source_ref,
                        "order_qty": order_qty,
                        "material_available": material_available,
                    })
                
        wb.close()
        return all_rows
        
    def get_order_qty(self, source_ref, path="Demo_ERP_table.xlsx") -> int | None:
        self._ensure_demo_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active
        assert ws is not None #to satisfy pylance

        for row in ws.iter_rows(min_row=2):
            cell_source_ref = row[0].value

            if str(cell_source_ref) == str(source_ref):
                value = row[1].value  # order_qty

                if isinstance(value, int):
                    wb.close()
                    return int(value)
                
                else: 
                    raise ValueError(f"order_qty: {value} is not INT")
        
        wb.close()
        return None

    def _ensure_demo_erp_exists(self, path="Demo_ERP_table.xlsx") -> None:
        """Create the demo ERP table if it does not exist."""
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "source_ref"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)
        wb.close()

# import custom_backends.py will override demo backends
try: from custom_backends import build_custom_backends # type: ignore
except ImportError: build_custom_backends = None


# ============================================================
# JOB FLOWS
# ============================================================

class MailFlow:
    """Handle email-driven job intake."""
    def __init__(self, logger, friends_repo, audit, is_within_operating_hours, network_service, personal_mail_handlers, shared_mail_handlers, job_lifecycle, personal_mailbox, shared_mailbox) -> None:
        self.logger = logger
        self.friends_repo = friends_repo
        self.audit = audit
        self._is_within_operating_schedule = is_within_operating_hours
        self.network_service = network_service
        self.personal_mail_handlers = personal_mail_handlers
        self.shared_mail_handlers = shared_mail_handlers
        self.job_lifecycle = job_lifecycle
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.friends_filename = Path(self.friends_repo.friends_path).name

    def poll_once(self) -> bool:
        if self._poll_personal_once():
            return True
        
        if self._poll_shared_once():
            return True
        
        return False

    def _poll_personal_once(self) -> bool:
        ''' direct human-to-robot channel (parse, always claim) '''

        personal_inbox_paths = self.personal_mailbox.list_inbox_mail_paths()

        if personal_inbox_paths:
            self.friends_repo.reload_if_modified()

        for path in personal_inbox_paths:
            active_job = self.personal_mailbox.parse_mail_file(path)

            if self.personal_mailbox._has_status_prefix(active_job, "DONE"):
                continue
            if self.personal_mailbox._has_status_prefix(active_job, "FAIL"):
                continue
            
            if self.personal_mailbox._has_status_prefix(active_job, "PROCESSING"):
                raise RobotRuntimeFault(message="stale personal mail found with subject prefix", phase="poll_intake", error_code="PRE_HANDOVER_CRASH", active_job=active_job)
            
            self._handle_personal_mail(active_job)
            return True

        return False
    
    def _poll_shared_once(self) -> bool:
        """External business mailbox: parse, skip irrelevant, handle one in-scope mail."""

        if not self._is_within_operating_schedule():
            return False

        if not self.network_service.has_network_access():
            return False

        shared_inbox_paths = self.shared_mailbox.list_inbox_mail_paths()

        for path in shared_inbox_paths:
            active_job = self.shared_mailbox.parse_mail_file(path)

            # already handled shared mail
            if self.shared_mailbox._has_status_prefix(active_job, "DONE"):
                continue
            if self.shared_mailbox._has_status_prefix(active_job, "FAIL"):
                continue
            if self.shared_mailbox._has_status_prefix(active_job, "PROCESSING"):
                self.logger.system(f"WARN: stale shared mail found with source_ref={active_job.source_ref}")
                continue

            # not our job
            handler = self._find_shared_mail_handler(active_job)
            if handler is None:
                continue
          
            self._handle_shared_mail(active_job, handler)
            return True

        return False

    def _handle_personal_mail(self, active_job: ActiveJob) -> None:
        result: PrecheckResult
        phase: RuntimePhase = "poll_intake"

        self.logger.system(f"{active_job.source_type} produced mail {active_job.source_ref}")
        self.logger.ui(f"email from {active_job.email_address}", blank_line_before=True)

        if not self.friends_repo.is_allowed_sender(active_job.email_address):
            self.job_lifecycle.delete_only(
                active_job,
                ui_log=f"--> rejected (not in {self.friends_filename})",
                system_log=f"--> rejected (not in {self.friends_filename})",
            )
            return
        
        # from this point, active_job is from a trusted user
        # (will enjoy a reply, audit logging, etc.)
        try:

            if not self._is_within_operating_schedule():
                self.job_lifecycle.reject_personal_mail(
                    active_job,
                    error_code="OUTSIDE_WORKING_HOURS",
                    ui_log="--> rejected (outside working hours)",
                )
                return
            
            handler = self._find_personal_mail_handler(active_job)
            if handler is None:
                self.job_lifecycle.reject_personal_mail(
                    active_job,
                    error_code="UNKNOWN_JOB",
                    ui_log="--> rejected (unable to identify job)",
                )
                return

            active_job.job_name = handler.job_name

            if not self.friends_repo.has_job_access(active_job.email_address, active_job.job_name):
                self.job_lifecycle.reject_personal_mail(
                    active_job,
                    error_code="NO_ACCESS",
                    ui_log=f"--> rejected (no access to {active_job.job_name})",
                )
                return

            if not self.network_service.has_network_access():
                self.job_lifecycle.reject_personal_mail(
                    active_job,
                    error_code="NO_NETWORK",
                    ui_log="--> rejected (no network connection)",
                )
                return
            
            phase = "personal_precheck"
            result = handler.precheck_and_build_payload(active_job)

            if not result.is_success:
                self.job_lifecycle.reject_personal_mail(
                    active_job,
                    error_code=result.error_code or "INVALID_INPUT",
                    public_error_message=result.public_error_message,
                    ui_log=f"--> rejected (invalid input for {active_job.job_name})",
                )
                return

            active_job.rpatool_payload = result.rpatool_payload
            active_job.request_summary = result.request_summary

            self.job_lifecycle.queue_for_rpa(
                active_job=active_job,
                send_online_notice=True,
                start_recording=True,
            )

        except RobotRuntimeFault:
            raise
            
        except Exception as err:
            self.logger.system(f"WARN: {err}", active_job.job_id)

            try:
                self.job_lifecycle.skip_due_to_pre_handover_crash(
                    active_job=active_job,
                    phase=phase,
                    internal_reason=str(err),
                    ui_log=f"--> fail ({active_job.job_name})",
                    traceback_text=traceback.format_exc(),
                )
                return
            
            # skip_due_to_pre_handover_crash() is designed to raise only RobotRuntimeFault      
            except RobotRuntimeFault:
                raise
                
    def _find_personal_mail_handler(self, active_job: ActiveJob):
        for handler in self.personal_mail_handlers.values():
            if handler.can_handle(active_job):
                return handler
        return None

    def _handle_shared_mail(self, active_job: ActiveJob, handler) -> None:
        result: PrecheckResult
        phase: RuntimePhase = "shared_precheck"

        self.logger.system("shared inbox mail detected (sender/subject omitted)")
        self.logger.ui(f"shared inbox email from {active_job.email_address}", blank_line_before=True,)

        active_job.job_name = handler.job_name

        try:
            result = handler.precheck_and_build_payload(active_job)

            if not result.is_success:
                self.job_lifecycle.skip_shared_mail(
                    active_job=active_job,
                    error_code=result.error_code or "INVALID_INPUT",
                    reason=result.public_error_message or "unable to handle input.",
                    ui_log=f"--> fail ({active_job.job_name})",
                )
                return

            active_job.rpatool_payload = result.rpatool_payload
            active_job.request_summary = result.request_summary

            self.job_lifecycle.queue_for_rpa(
                active_job=active_job,
                send_online_notice=False,
                start_recording=True,
            )
            return
        
        except RobotRuntimeFault:
            raise
            
        except Exception as err:
            self.logger.system(f"WARN: {err}", active_job.job_id)

            try:
                self.job_lifecycle.skip_due_to_pre_handover_crash(
                    active_job=active_job,
                    phase=phase,
                    internal_reason=str(err),
                    ui_log=f"--> fail ({active_job.job_name})",
                    traceback_text=traceback.format_exc(),
                )
                return
            
            # skip_due_to_pre_handover_crash() is designed to raise only RobotRuntimeFault      
            except RobotRuntimeFault:
                raise
      
    def _find_shared_mail_handler(self, active_job: ActiveJob):
        for handler in self.shared_mail_handlers.values():
            if handler.can_handle(active_job):
                return handler
        return None


class QueryFlow:
    """Handle query-driven job intake."""

    def __init__(self, logger, query_handlers, audit, job_lifecycle, is_within_operating_hours) -> None:
        self.logger = logger
        self.query_handlers = query_handlers
        self.audit = audit
        self.job_lifecycle = job_lifecycle
        self._is_within_operating_schedule = is_within_operating_hours

        self.logged_skipped_date = datetime.date.today()
        self.logged_skipped_today: set[str] = set()

    def poll_once(self) -> bool:
        # intentionally crash the robot if query source broken w/o try-block

        self._set_skipped_jobs_date()

        if not self._is_within_operating_schedule():
            return False

        for handler in self.query_handlers.values():
            active_job: ActiveJob | None = None
            #self.logger.system(f"checking query handler {handler.job_name}")

            active_job_candidates = handler.find_next_active_jobs()
            self._validate_candidate(active_job_candidates, handler)
            
            for candidate in active_job_candidates:
      
                if self.audit.has_been_processed_today(candidate.source_ref):
                    if candidate.source_ref not in self.logged_skipped_today:
                        self.logger.system(f"skip {candidate.source_ref} (already handled today)")
                        self.logged_skipped_today.add(candidate.source_ref)
                    continue

                active_job = candidate
                break

            if active_job is None:
                continue

            self._handle_query_row(active_job, handler)
            return True
           
        return False

    def _handle_query_row(self, active_job: ActiveJob, handler) -> None:
        phase: RuntimePhase = "poll_intake"
        result: PrecheckResult   

        self.logger.ui(f"query job detected: {active_job.source_ref}", blank_line_before=True,)
        phase = "query_precheck"
        
        try:
            result = handler.precheck_and_build_payload(active_job) 

            if not result.is_success:
                error_code = "INVALID_INPUT"
                self.logger.system(f"query candidate rejected, source_ref={active_job.source_ref}, error_message={result.public_error_message}")
                self.job_lifecycle.reject_query_result(
                active_job=active_job,
                error_code=result.error_code or error_code,
                public_error_message=result.public_error_message,
                ui_log=f"--> fail ({active_job.job_name})",
                )
                return

            active_job.rpatool_payload = result.rpatool_payload
            active_job.request_summary = result.request_summary

            self.job_lifecycle.queue_for_rpa(
                active_job=active_job,
                send_online_notice=False,
                start_recording=True,
                # ui_log=f"--> accepted ({handler.job_name})",
            )
            return
        
        except RobotRuntimeFault:
            raise
            
        except Exception as err:
            self.logger.system(f"WARN: {err}", active_job.job_id)

            try:
                self.job_lifecycle.skip_due_to_pre_handover_crash(
                    active_job=active_job,
                    phase=phase,
                    internal_reason=str(err),
                    ui_log=f"--> fail ({active_job.job_name})",
                    traceback_text=traceback.format_exc(),
                )
                return
            
            # skip_due_to_pre_handover_crash() is designed to raise only RobotRuntimeFault      
            except RobotRuntimeFault:
                raise

    def _validate_candidate(self, candidate_jobs, handler) -> None:
        
        for candidate in candidate_jobs:
            if not hasattr(candidate, "source_ref") :
                raise ValueError(f"{handler.job_name}-handler returned candidate without .source_ref")

            if not hasattr(candidate, "source_type"):
                raise ValueError(f"{handler.job_name}-handler returned candidate without .source_type")

            if candidate.source_type != "erp_query":
                raise ValueError(
                    f"{handler.job_name}-handler should return source_type=erp_query, got {candidate.source_type}"
                )
            
            if not isinstance(candidate, ActiveJob):
                raise ValueError(
                f"{handler.job_name}-handler should return ActiveJob objects, got {type(candidate)}"
            )
        
    def _set_skipped_jobs_date(self) -> None:
        today = datetime.date.today()
        if today != self.logged_skipped_date:
            self.logged_skipped_today.clear()
            self.logged_skipped_date = today

# ============================================================
# JOB HANDLERS
# ============================================================

class PingHandler:
    '''This 'automation' allows the user to check if the robot (RobotRuntime + RPA tool) running.'''

    job_name: JobName = "ping"

    def __init__(self, logger) -> None:
        self.logger = logger

    def can_handle(self, active_job: ActiveJob) -> bool:
            subject = str(active_job.email_subject).strip()
            original_subject = re.sub(r"^(PROCESSING|DONE|FAIL)/\d{8}/", "", subject, flags=re.IGNORECASE).strip().lower()
            return original_subject == self.job_name

    def precheck_and_build_payload(self, active_job: ActiveJob) -> PrecheckResult:
        """Validate the request and build the payload for the RPA tool."""

        
        return PrecheckResult(is_success=True, rpatool_payload={})

    def verify_result(self, active_job: ActiveJob) -> VerificationResult:
        '''
        verify_result() must return:
        - success, or
        - failure with error_code=POST_HANDOVER_... and public_error_message
        
        Other outcomes are treated as programming/system faults by RobotRuntime. 
        '''
        return VerificationResult(is_success=True)

try: from custom_query_jobs import build_custom_query_handlers # type: ignore
except ImportError: build_custom_query_handlers = None

try: from custom_personal_mail_jobs import build_custom_personal_mail_handlers # type: ignore
except ImportError: build_custom_personal_mail_handlers = None

try: from custom_shared_mail_jobs import build_custom_shared_mail_handlers # type: ignore
except ImportError: build_custom_shared_mail_handlers = None


# ============================================================
# HANDOVER
# ============================================================

class HandoverRepository:
    """Persist and validate the file-based state shared with the RPA tool."""

    def __init__(self, logger, handover_file) -> None:
        self.logger = logger
        self.handover_file = handover_file

    def read(self) -> HandoverFile:
        ''' read HANDOVER_FILE '''
        
        last_err=None

        for attempt in range(7):
            try:
                # read file
                with open(self.handover_file, "r", encoding="utf-8") as f:
                    handover_data = json.load(f)
                
                # rebuild object
                handover_file = self._validate_and_build_handover_file(handover_data)

                return handover_file
                
            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: retry {attempt+1}/7 : {err}")
                time.sleep(attempt/10)
        
        
        raise RuntimeError(f"{self.handover_file} unreadable: {last_err}")    
      
    def write(self, handover_file: HandoverFile) -> None:
        ''' atomic write of HANDOVER_FILE '''

        handover_data = asdict(handover_file)

        self._validate_and_build_handover_file(handover_data) # only validate (ignore return)
        job_id = handover_data.get("job_id")

        last_err = None
        
        for attempt in range(7):
            temp_path = None
            try:
                
                dir_path = os.path.dirname(os.path.abspath(self.handover_file))
                fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

                #atomic write
                with os.fdopen(fd, "w", encoding="utf-8") as tmp:
                    json.dump(handover_data, tmp, indent=2) # indent for human eyes
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(temp_path, self.handover_file)
                
                self.logger.system(
                    f"wrote handover state={handover_data.get('state')}, job_name={handover_data.get('job_name')}, "
                    f"rpatool_payload={handover_data.get('rpatool_payload')} etc. (GDPR sanitized)",
                    job_id,
                )               
                return

            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: {attempt+1}/7 error", job_id)
                time.sleep(attempt/10) # 0 0.1... 0.6 sec     

            finally:
                if temp_path and os.path.exists(temp_path):
                    try: os.remove(temp_path)
                    except Exception as err: self.logger.system(f"{err}")

        self.logger.system(f"CRITICAL: cannot write {self.handover_file} {last_err}", job_id)
        raise RuntimeError(f"CRITICAL: cannot write {self.handover_file}")

    def is_valid_observed_transition(self, prev_state: HandoverState | None, state: HandoverState) -> bool:
        """Validate transitions observed by polling. Allows skipped states."""

        if prev_state is None: # at startup
            return True

        allowed_observed: dict[HandoverState, set[HandoverState]] = {
            "idle": {"job_queued", "job_running", "job_verifying", "safestop"},
            "job_queued": {"job_running", "job_verifying", "safestop"},
            "job_running": {"job_verifying", "safestop"},
            "job_verifying": {"idle", "safestop"},
            "safestop": {"idle"},
        }

        allowed_next = allowed_observed[prev_state]

        return state in allowed_next
       
    def _validate_and_build_handover_file(self, handover_data: dict) -> HandoverFile:
        state = handover_data.get("state")
        job_id = handover_data.get("job_id")
        job_name = handover_data.get("job_name")
        rpatool_payload = handover_data.get("rpatool_payload")

        if state is None:
            raise ValueError("state missing")

        if state not in get_args(HandoverState):
            raise ValueError(f"unknown state: {state}")

        if job_id is not None:
            job_id = int(job_id)

        if state == "idle":
            if any(v is not None for v in (job_id, job_name, rpatool_payload)):
                raise ValueError(f"state 'idle' should have no job data: {handover_data}")

        elif state in ("job_queued", "job_running", "job_verifying"):
            if job_id is None:
                raise ValueError(f"{state} missing job_id")
            if not job_name:
                raise ValueError(f"{state} missing job_name")
            if rpatool_payload is None:
                raise ValueError(f"{state} missing rpatool_payload")
            if not isinstance(rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")

        elif state == "safestop":
            pass

        return HandoverFile(
            state=state,
            job_id=job_id,
            job_name=job_name,
            rpatool_payload=rpatool_payload,
        )


class JobLifecycleService:
    """Execute pre-handover actions and build HandoverFile objects for the RPA tool."""

    def __init__(self, logger, handover, show_recording_overlay, recording, audit, notifications, personal_mailbox, shared_mailbox, job_handlers, hide_recording_overlay, generate_job_id) -> None:
        self.logger = logger
        self.handover = handover
        self.recording = recording
        self.audit = audit
        self._show_recording_overlay = show_recording_overlay
        self.notifications = notifications
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.job_handlers = job_handlers
        self._hide_recording_overlay = hide_recording_overlay
        self.generate_job_id = generate_job_id

    def delete_only(self, active_job: ActiveJob, ui_log: str, system_log: str | None = None) -> None:
        if active_job.job_id is not None:
            raise ValueError("delete_only must not be used for audit-tracked jobs")
        
        self.logger.ui(ui_log)

        if system_log:
            self.logger.system(system_log)

        self.personal_mailbox.delete(active_job, fallback_status="DONE")

    def reject_personal_mail(self, active_job: ActiveJob, error_code: LifecycleErrorCode, public_error_message: str | None = None, ui_log: str | None = None,) -> None:
        lifecycle_status: LifecycleStatus = "REJECTED"
        
        if ui_log:
            self.logger.ui(ui_log)

        active_job.job_id = self.generate_job_id()

        self.audit.insert(
            active_job=active_job,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            error_message=public_error_message,
        )

        self.notifications.send_final_reply(
            active_job=active_job,
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            public_error_message=public_error_message,
        )

        self.close_personal_mail_after_final_reply(
            active_job=active_job,
            fallback_status="DONE",
        )

    def reject_query_result(self, active_job: ActiveJob, error_code: LifecycleErrorCode, public_error_message: str | None = None, ui_log: str | None = None,) -> None:
        lifecycle_status: LifecycleStatus = "REJECTED"
        
        if ui_log:
            self.logger.ui(ui_log)

        active_job.job_id = self.generate_job_id()

        self.audit.insert(
            active_job=active_job,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            error_message=public_error_message,
        )

    def skip_shared_mail(self, active_job: ActiveJob, error_code: LifecycleErrorCode, reason: str, ui_log: str | None = None,) -> None:
        lifecycle_status: LifecycleStatus = "FAIL"
        
        if active_job.job_id is not None:
            raise ValueError("skip_shared_mail should only be called in early lifecycle")
        
        self.logger.system("running")

        if ui_log:
            self.logger.ui(ui_log)
        
        active_job.job_id = self.generate_job_id()

        if not active_job.job_name:
            active_job.job_name = "unknown"

        self.audit.insert(
            active_job=active_job,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            error_message=reason,
            final_reply_sent=False,
        )

        self.shared_mailbox.mark_failed(active_job)

    def queue_for_rpa(self, active_job: ActiveJob, send_online_notice: bool, start_recording: bool, ui_log: str | None = None,) -> None:
        lifecycle_status: LifecycleStatus = "QUEUED"
        phase: RuntimePhase = "queue_for_rpa"

        try:
            self.logger.system("running")
            
            if ui_log:
                self.logger.ui(ui_log)

            if not active_job.job_name:
                raise ValueError("active_job missing .job_name")
            
            if active_job.rpatool_payload is None:
                raise ValueError("active_job missing .rpatool_payload, need atleast empty dict")

            active_job.job_id = self.generate_job_id()

            if active_job.source_type == "personal_inbox":
                active_job = self.personal_mailbox.mark_processing(active_job)

            elif active_job.source_type == "shared_inbox":
                active_job = self.shared_mailbox.mark_processing(active_job)

            elif active_job.source_type == "erp_query":
                pass

            else:
                raise ValueError(f"unknown active_job.source_type={active_job.source_type}")

            should_send_online_notice = (send_online_notice and not self.audit.has_sender_job_today(active_job))

            self.audit.insert(
                active_job=active_job,
                started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
                started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
                lifecycle_status=lifecycle_status,
            )
            
            if start_recording:
                self._maybe_start_recording(active_job)
            
            if should_send_online_notice:
                try: self.notifications.send_online_notice(active_job)
                except Exception as notice_err: self.logger.system( f"WARN: nice-to-have online notice failed with error={notice_err}", active_job.job_id,) 

            handover = HandoverFile(
                state="job_queued",
                job_name=active_job.job_name,
                job_id=active_job.job_id,
                rpatool_payload=active_job.rpatool_payload,
            )
            self.handover.write(handover)

        except Exception as err:  
            try:
                self.skip_due_to_pre_handover_crash(
                    active_job=active_job,
                    phase=phase,
                    internal_reason=str(err),
                    ui_log=f"--> fail ({active_job.job_name})",
                    traceback_text=traceback.format_exc(),
                )
                return
            
            # skip_due_to_pre_handover_crash() is designed to raise only RobotRuntimeFault      
            except RobotRuntimeFault:
                raise     

    def skip_due_to_pre_handover_crash(self, active_job: ActiveJob, phase: RuntimePhase, internal_reason: str, ui_log: str | None = None, traceback_text: str | None = None,) -> None:
        """No damage in ERP before handover.json is written -> fail/skip this job and keep robot alive"""
        
        try:
            lifecycle_status: LifecycleStatus = "FAIL"
            error_code: LifecycleErrorCode = "PRE_HANDOVER_CRASH"

            self.logger.system("running", active_job.job_id)
            
            if active_job.job_id:
                self._hide_recording_overlay()
                self.recording.stop(active_job.job_id)
                self.recording.try_upload_recording(active_job.job_id)

            if ui_log:
                self.logger.ui(ui_log)

            if active_job.job_id is None:
                active_job.job_id = self.generate_job_id()

            if not active_job.job_name:
                active_job.job_name = "unknown"

            existing = self.audit.get_row_by_id(active_job.job_id)

            if existing:
                self.audit.mark_failed(
                    job_id=active_job.job_id,
                    error_code=error_code,
                    error_message=internal_reason,
                )
            else:
                self.audit.insert(
                    active_job=active_job,
                    started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
                    started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
                    lifecycle_status=lifecycle_status,
                    error_code=error_code,
                    error_message=internal_reason,
                )
        
            if active_job.source_type == "personal_inbox":
                if self.audit.final_reply_sent(active_job.job_id) or self.personal_mailbox.sent_reply_exists(active_job.source_ref):
                    self.logger.system("skip send_final_reply since already sent", active_job.job_id)
                        
                else:
                    self.notifications.send_final_reply(
                        active_job=active_job,
                        lifecycle_status=lifecycle_status,
                        error_code=error_code,
                    )

                self.close_personal_mail_after_final_reply(
                    active_job=active_job,
                    fallback_status="FAIL",
                )
                
            elif active_job.source_type == "shared_inbox":
                self.shared_mailbox.mark_failed(active_job)
            
            elif active_job.source_type == "erp_query":
                # audit row is enough, assuming every query handler use has_been_processed_today(source_ref) to block bad rows
                pass

            else:
                raise ValueError(f"unknown active_job.source_type={active_job.source_type}")

            # exception in admin alert will intentionally crash the robot
            self.notifications.send_admin_alert(
                reason=(
                    f"error_code={error_code}\n\n"
                    f"phase={phase}\n"
                    f"source_type={active_job.source_type}\n"
                    f"job_name={active_job.job_name}\n"
                    f"source_ref={active_job.source_ref}\n"
                    f"error={internal_reason}\n\n"
                    f"{traceback_text or 'Traceback unavailable'}"
                    ),
                critical=False,
                )
        
        except Exception as recovery_err:
            self.logger.system(f"CRIT: recovery_err {recovery_err}", active_job.job_id)

            raise RobotRuntimeFault(
                message=f"double-crash in {active_job.job_name}",
                active_job=active_job,
                phase=phase,
                error_code=error_code,
                cause=recovery_err,
                traceback_text=traceback.format_exc(),
            ) from recovery_err
        
    def complete_from_handover(self, handover_file: HandoverFile) -> None:
        active_job: ActiveJob
        jobhandler_verification_result: VerificationResult | None = None
        phase: RuntimePhase = "verification"
        error_code: LifecycleErrorCode = "POST_HANDOVER_UNSPEC_CRASH"

        active_job = self.audit.parse_from_jobaudit(handover_file.job_id)
        self.logger.system(f"completing {active_job.job_name} with payload {active_job.rpatool_payload}", active_job.job_id) # only store safe data in log
       
        try:    
            self.audit.mark_verifying(active_job.job_id)

            self._hide_recording_overlay()
            self.recording.stop(active_job.job_id)
            self.recording.try_upload_recording(active_job.job_id)
        
            handler = self.job_handlers.get(active_job.job_name)
            if handler == None:
                raise ValueError(f"handler missing for job_name={active_job.job_name}")

            jobhandler_verification_result = handler.verify_result(active_job)
            if jobhandler_verification_result is None:
                raise ValueError(f"handler for job_name={active_job.job_name} returned no result")
            
            self._validate_format(jobhandler_verification_result)

            if jobhandler_verification_result.is_success:
                self._complete_successful_result(active_job)
                self._update_logs("DONE", active_job)
                self.handover.write(HandoverFile(state="idle"))
                return
            
            self._complete_failed_result(active_job, jobhandler_verification_result)
            self._update_logs("FAIL", active_job)
            self.handover.write(HandoverFile(state="idle"))

            # policy to safestop if post-handover error    
            raise RobotRuntimeFault(
                message=jobhandler_verification_result.public_error_message or "verification failed",
                phase=phase,
                error_code=jobhandler_verification_result.error_code or error_code,
                handover_file=handover_file,
                active_job=active_job,
            )

        except RobotRuntimeFault:
            raise

        except Exception as err:

            if jobhandler_verification_result is not None and jobhandler_verification_result.error_code is not None:
                error_code = jobhandler_verification_result.error_code

            try:
                existing = self.audit.get_row_by_id(active_job.job_id)
                if existing and existing.get("lifecycle_status") != "DONE":
                    self.audit.mark_failed(
                        job_id=active_job.job_id,
                        error_code=error_code,
                        error_message=f"crash during verification stage",
                    )
            except Exception as err2:
                self.logger.system(f"double-error {err} {err2}", active_job.job_id)

            raise RobotRuntimeFault(
                #str(err),
                phase=phase,
                error_code=error_code,
                handover_file=handover_file,
                active_job=active_job,
                cause=err,
            ) from err

    def close_personal_mail_after_final_reply(self, active_job: ActiveJob, fallback_status: Literal["DONE", "FAIL"]) -> None:
        if active_job.job_id is None:
            raise ValueError("active_job.job_id is None")

        try:
            self.audit.mark_final_reply_sent(active_job.job_id)

        except Exception as err:
            self.logger.system(
                f"error in .mark_final_reply_sent, fallback to .mark_done/.mark_failed {fallback_status} and keeping it, error={err}",
                active_job.job_id,
            )

            if fallback_status == "DONE":
                self.personal_mailbox.mark_done(active_job)
            else:
                self.personal_mailbox.mark_failed(active_job)

            return  # dont delete

        self.personal_mailbox.delete(active_job, fallback_status=fallback_status)
         
    def _complete_successful_result(self, active_job: ActiveJob) -> None:
        lifecycle_status: LifecycleStatus = "DONE"
        self.audit.mark_done(active_job.job_id)

        if active_job.source_type == "personal_inbox":
            self.notifications.send_final_reply(
                active_job=active_job,
                lifecycle_status=lifecycle_status,
            )
            self.close_personal_mail_after_final_reply(
                active_job=active_job,
                fallback_status="DONE",
            )

        elif active_job.source_type == "shared_inbox":
            self.shared_mailbox.mark_done(active_job)
        
        elif active_job.source_type == "erp_query":
            pass

        else:
            raise ValueError(f"unknown active_job.source_type={active_job.source_type}")

    def _complete_failed_result(self, active_job: ActiveJob, jobhandler_verification_result: VerificationResult) -> None:
        lifecycle_status: LifecycleStatus = "FAIL"
                                                   
        self.audit.mark_failed(
            job_id=active_job.job_id, 
            error_code=jobhandler_verification_result.error_code, 
            error_message=jobhandler_verification_result.public_error_message
            )

        if active_job.source_type == "personal_inbox":
            self.notifications.send_final_reply(
                active_job=active_job,
                lifecycle_status=lifecycle_status,
                error_code=jobhandler_verification_result.error_code,
                public_error_message=jobhandler_verification_result.public_error_message,
                recovery_context="safestop",
            )

            self.close_personal_mail_after_final_reply(
                active_job=active_job,
                fallback_status="FAIL",
            )
            return
            
        if active_job.source_type == "shared_inbox":
            self.shared_mailbox.mark_failed(active_job)
            return

        if active_job.source_type == "erp_query":
            return
        
        raise ValueError(f"unknown active_job.source_type={active_job.source_type}")
         
    def _maybe_start_recording(self, active_job: ActiveJob):
        started = self.recording.start(active_job)
        if started:
            try: self._show_recording_overlay()
            except Exception as e: self.logger.system(f"error {e}", active_job.job_id)

    def _validate_format(self, jobhandler_verification_result: VerificationResult):
        allowed_error_codes:list[LifecycleErrorCode] = ["POST_HANDOVER_VERIFICATION_MISMATCH", "POST_HANDOVER_VERIFICATION_TIMEOUT", "POST_HANDOVER_UNSPEC_CRASH"]
        
        if jobhandler_verification_result.is_success:
            if jobhandler_verification_result.error_code:
                raise ValueError(f"error_code must be empty for is_success=True, is {jobhandler_verification_result.error_code}")
            if jobhandler_verification_result.public_error_message:
                raise ValueError(f"error_message must be empty for is_success=True, is {jobhandler_verification_result.public_error_message}")
        else:   
            if jobhandler_verification_result.error_code not in allowed_error_codes:
                raise ValueError(f"jobhandler_verification_result.error_code={jobhandler_verification_result.error_code} from handler is not in allowed {allowed_error_codes}")
            if not jobhandler_verification_result.public_error_message:
                raise ValueError("missing verification_result.public_error_message for is_success=False")
        
    def _update_logs(self, lifecycle_status: Literal["DONE", "FAIL"], active_job: ActiveJob,) -> None:
        self.logger.ui(f"--> {lifecycle_status.lower()} ({active_job.job_name})")
        self.logger.system(f"{lifecycle_status} ({active_job.job_name})", active_job.job_id)
   
     
# ============================================================
# USER NOTIFICATIONS
# ============================================================

class UserNotificationService:
    """Only for personal_inbox user replies."""

    def __init__(self, logger, personal_mailbox, friends_repo, config: RuntimeConfig):
        self.logger = logger
        self.personal_mailbox = personal_mailbox
        self.friends_repo = friends_repo
        self.config = config
    
    def send_final_reply(self, active_job: ActiveJob, lifecycle_status: LifecycleStatus, error_code: LifecycleErrorCode | None=None, recovery_context: str | None=None, public_error_message: str | None = None,) -> None:
        if active_job.job_id == None:
            raise ValueError("active_job.job_id is None")

   
        extra_subject, extra_body = self._build_reply(
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            job_id=active_job.job_id,
            public_error_message=public_error_message,
            recovery_context=recovery_context,
            request_summary=active_job.request_summary,
        )

        self._reply(active_job, extra_subject, extra_body)

    def send_out_of_service_reply(self, active_job: ActiveJob) -> None:
        lifecycle_status: LifecycleStatus = "FAIL"
        error_code: LifecycleErrorCode = "OUT_OF_SERVICE"

        if active_job.job_id is None:
            raise ValueError("send_out_of_service_reply requires active_job.job_id to be assigned by lifecycle/recovery")
        
        self.send_final_reply(
            active_job=active_job,
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            )

    def send_command_reply(self, active_job: ActiveJob) -> None:

        self._reply(
            active_job=active_job,
            extra_subject="got it!",
            extra_body="Command received.\n" + self._get_robot_signature(),
        )
  
    def send_admin_alert(self, reason: str, critical: bool = False) -> None:
        fake_active_job = ActiveJob(
            source_ref="robot",
            job_id=self.config.mail_command_job_id,
            email_address=self.config.rpa_admin_email,
            email_subject="",
            email_body="",
            source_type="personal_inbox",
            source_data={},
        )

        body = (
                f"Robot encountered an error.\n"
                f"Reason:\n{reason}\n\n"
        )
        
        if critical:
            body += (
                "Robot is now in degraded mode.\n\n"
                "Reminder of available email commands: 'stop1234' and 'restart1234'."
            )
        else:
            body += (
                "The error was non-critical and robot is still running"
            )

        self._reply(
            active_job=fake_active_job,
            extra_subject="CRIT" if critical else "WARN",
            extra_body=body,
        )

    def send_online_notice(self, active_job: ActiveJob) -> None:
        # TODO add average completion time calculated from job_audit for requested job_name 
      
        body = (
            ">Hello, human<\n\n"
            "The first request each day is replied with: online\n"
            "You should receive a final reply after completion\n"
            f"(in max {(self.config.rpa_tool_claim_timeout + self.config.rpa_tool_execution_timeout + 59) // 60} min from now)."
        )
    
        body += self._get_robot_signature()

        self._reply(
            active_job=active_job,
            extra_subject="ONLINE",
            extra_body=body,
        )

    def _classify_reply_kind(self, lifecycle_status: LifecycleStatus, error_code: LifecycleErrorCode | None) -> str:

        if lifecycle_status == "DONE":
            return "DONE"
        
        if lifecycle_status == "FAIL" and error_code == "PRE_HANDOVER_CRASH":
            return "NOT_STARTED"

        if lifecycle_status == "FAIL" and error_code == "OUT_OF_SERVICE":
            return "OUT_OF_SERVICE"

        if lifecycle_status == "FAIL" and error_code == "RPA_TOOL_CRASH":
            return "STARTED_BUT_CRASHED"

        if lifecycle_status == "FAIL" and error_code == "POST_HANDOVER_VERIFICATION_MISMATCH":
            return "POST_HANDOVER_VERIFICATION_MISMATCH"
        
        if lifecycle_status == "FAIL" and error_code == "POST_HANDOVER_VERIFICATION_TIMEOUT":
            return "VERIFYING_CRASH"

        if lifecycle_status == "FAIL" and error_code == "POST_HANDOVER_UNSPEC_CRASH":
            return "VERIFYING_CRASH"

        if lifecycle_status == "VERIFYING":
            return "VERIFYING_CRASH"
        
        if lifecycle_status == "REJECTED":
            return "NOT_STARTED"

        if lifecycle_status == "QUEUED":
            return "NOT_STARTED"

        if lifecycle_status == "RUNNING":
            return "STARTED_BUT_CRASHED"

        if lifecycle_status == "FAIL":
            return "UNKNOWN_FAIL"

        raise ValueError(f"Cannot classify reply for lifecycle_status={lifecycle_status}, error_code={error_code}")

    def _get_standard_reason(self, error_code: LifecycleErrorCode | None) -> str | None:
        if error_code == "OUTSIDE_WORKING_HOURS":
            return (
                "Outside robot's working hours "
                f"{self.config.operating_hours_start}:00-{self.config.operating_hours_end}:00."
            )

        if error_code == "NO_NETWORK":
            return "No network connection at the moment."

        if error_code == "UNKNOWN_JOB":
            return "Could not identify a job type from your email."

        if error_code == "NO_ACCESS":
            return "Request denied. Your email is not permitted to trigger this job."

        if error_code == "IN_SAFESTOP":
            return "Robot is currently out-of-service."
        
        if error_code == "PRE_HANDOVER_CRASH":
            return "unknown crash"

        return None

    def _build_reply(self, lifecycle_status: LifecycleStatus, job_id: int, error_code: LifecycleErrorCode | None, recovery_context: str | None=None, public_error_message: str | None = None, request_summary: str | None=None) -> tuple[str, str]:

        subject: str
        body: str

        recording_text = self._get_recording_text(job_id)
        reply_kind = self._classify_reply_kind(lifecycle_status, error_code)
        standard_reason = self._get_standard_reason(error_code)
        admin_text = self._get_admin_text(error_code)
        
        public_reason = public_error_message or standard_reason

        if reply_kind == "DONE":
            subject = "DONE"
            body = (
                    f"Job completed successfully.\n\n"
                    #f"job_id: {job_id}\n\n"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    f"{recording_text}"
                    f"No need to save this email for future reference.\n"
                )

        elif reply_kind == "NOT_STARTED":
            subject = "FAIL"
            body = (
                    f"Your request was not started.\n\n"
                    f"{f'Reason: {public_reason}\n\n' if public_reason else ''}"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    f"No changes were made in {self.config.erp_system_name}.\n"
                    f'{admin_text}'
                    f"No need to save this email for future reference.\n"
            )
              

        elif reply_kind == "STARTED_BUT_CRASHED":
            subject = "FAIL"
            body = (
                    f"Robot started your request, but then crashed.\n\n"
                    f"{f'Reason: {public_reason}\n\n' if public_reason else ''}"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    #f"job_id: {job_id}\n"
                    f"The request is probably 'half completed' in {self.config.erp_system_name}.\n"
                    f"It is (very) recommended that you review the result manually.\n\n"
                    f"{recording_text}"
                    f'{admin_text}'
                    f"No need to save this email for future reference.\n"
                )
                
        elif reply_kind == "POST_HANDOVER_VERIFICATION_MISMATCH":
            subject = "FAIL"
            body = (
                    f"Robot completed the request, and the result was checked in {self.config.erp_system_name}.\n"
                    f"However, the final {self.config.erp_system_name} data did not match the expected result.\n\n"
                    f"{f'Reason: {public_reason}\n\n' if public_reason else ''}"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    #f"job_id: {job_id}\n"
                    f"It is (very) recommended that you review the result manually.\n\n"
                    f"{recording_text}"
                    f'{admin_text}'
                    f"No need to save this email for future reference.\n"
                )

        elif reply_kind == "VERIFYING_CRASH":
            subject = "FAIL"
            body = (
                    f"Robot completed the request in {self.config.erp_system_name}, but crashed during the final verification stage.\n"
                    f"The outcome could therefore not be confirmed automatically.\n\n"
                    #f"{f'Reason: {public_reason}\n\n' if public_reason else ''}"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    #f"job_id: {job_id}\n"
                    f"Please verify the result manually in {self.config.erp_system_name}.\n\n"
                    f"{recording_text}"
                    f'{admin_text}'
                    f"No need to save this email for future reference.\n"
                )

        elif reply_kind == "OUT_OF_SERVICE":
            subject = "FAIL"
            body = (
                    f"Robot is temporary out-of-service and does not accept any new requests.\n"
               )

        elif reply_kind == "UNKNOWN_FAIL":
            subject = "FAIL"
            body = (
                    f"Robot crashed and the exact job outcome could not be classified.\n\n"
                    f"{f'Reason: {public_reason}\n\n' if public_reason else ''}"
                    f"{f'Request summary:\n{request_summary}\n\n' if request_summary else ''}"
                    #f"job_id: {job_id}\n"
                    f"Please review the result manually in {self.config.erp_system_name}.\n\n"
                    f"{recording_text}"
                    f'{admin_text}'
                    f"No need to save this email for future reference.\n"
                )

        else:
            raise ValueError(f"Unhandled reply_kind={reply_kind}")
        

        if recovery_context == "safestop":
            body += "Robot will now temporary go out-of-service.\n"

        elif recovery_context == "startup":
            body = (
                    "Robot was offline and has now restarted.\n"
                    "If you already received a final reply (DONE/FAIL) for this job, you can ignore this recovery message.\n\n"
                ) + body


        body += self._get_robot_signature()
        body += f"tag {job_id}"

        return subject, body
  
    def _get_recording_text(self, job_id: int) -> str:
        recording_path = Path(self.config.recordings_destination_folder) / f"{job_id}.mp4"
        if recording_path.exists():
            return (
                "A screen recording is available for review:\n"
                f"{recording_path}\n\n"
            )

        return ""

    def _get_admin_text(self, error_code: LifecycleErrorCode | None) -> str:
        if error_code in {
            "PRE_HANDOVER_CRASH",
            "RPA_TOOL_CRASH",
            "POST_HANDOVER_VERIFICATION_MISMATCH",
            "POST_HANDOVER_VERIFICATION_TIMEOUT",
            "POST_HANDOVER_UNSPEC_CRASH",
            "RECOVERY_SOURCE_MISSING",
        }:
            return "A malfunction report will be sent to robot admin.\n\n"

        return "" 

    def _get_robot_signature(self) -> str:
        return (
            "\n---\n"
            "Automated message from Robot.\n"
        )

    def _reply(self, active_job: ActiveJob, extra_subject: str, extra_body: str,) -> None:

        if active_job.email_address is None:
            raise ValueError("cannot reply because active_job.email_address is None")
        
        # extra gatekeeper
        if not active_job.email_address == self.config.rpa_admin_email:
            if not self.friends_repo.is_allowed_sender(active_job.email_address):
                raise RuntimeError(f"CRIT: code-base trying to reply to email outside friends access list")

        self.personal_mailbox.send_reply(
            active_job=active_job,
            extra_subject=extra_subject,
            extra_body=extra_body,
        )

        self.logger.system(f"message sent with extra_body={extra_body[:100]}... (GDPR sanitized)", active_job.job_id)


# ============================================================
# RECORDING / SAFESTOP / INFRASTRUCTURE
# ============================================================   
                      
class RecordingService:
    ''' screen-recorder to capture all RPA tool screen-activity '''

    def __init__(self, logger, recordings_in_progress_folder, recordings_destination_folder) -> None:
        self.logger = logger
        self.recording_process = None
        self._ffmpeg_warned = False
        self.recordings_in_progress_folder = recordings_in_progress_folder
        self.recordings_destination_folder = recordings_destination_folder
 
    def start(self, active_job: ActiveJob) -> bool:
        """start the screen recording"""
        # written by AI
        try:
            job_id = active_job.job_id

            os.makedirs(self.recordings_in_progress_folder, exist_ok=True)
            filename = f"{self.recordings_in_progress_folder}/{job_id}.mp4"

            drawtext = (
                f"drawtext=text='job_id  {job_id}':"
                "x=200:y=20:"
                "fontsize=32:"
                "fontcolor=lightyellow:"
                "box=1:"
                "boxcolor=black@0.5"
            )

            if platform.system() == "Windows":
                ffmpeg_path = None

                local_ffmpeg = Path("./ffmpeg.exe")
                if local_ffmpeg.exists():
                    ffmpeg_path = str(local_ffmpeg)
                else:
                    ffmpeg_in_path = shutil.which("ffmpeg")
                    if ffmpeg_in_path:
                        ffmpeg_path = ffmpeg_in_path

                if ffmpeg_path is None:
                    self.logger.ui("--> recording disabled (ffmpeg missing)")
                    
                    if not self._ffmpeg_warned:
                        message = (
                            "FFMPEG.exe NOT FOUND\n\n"
                            "Screen recording is disabled.\n\n"
                            "Fix:\n"
                            "1. Go to: https://www.gyan.dev/ffmpeg/builds/\n"
                            "2. Download 'ffmpeg-git-essentials'\n"
                            "3. Extract the archive\n"
                            "4. Open the 'bin' folder\n"
                            "5. Copy ffmpeg.exe next to main.py\n"
                        )

                        print("\n" + "="*60 + "\n" + message + "\n" + "="*60 + "\n")
                        self.logger.system(message, job_id)
                        self._ffmpeg_warned = True
                        
                    return False

                capture = ["-f", "gdigrab", "-i", "desktop"]

                recording_process = subprocess.Popen(
                    [
                        ffmpeg_path,
                        "-y",
                        *capture,
                        "-framerate", "15",
                        "-vf", drawtext,
                        "-vcodec", "libx264",
                        "-pix_fmt", "yuv420p",
                        "-preset", "ultrafast",
                        filename,
                    ],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
                )

            else:
                display = os.environ.get("DISPLAY")
                if not display:
                    self.logger.system("WARN: screen-recording disabled because DISPLAY is missing", job_id)
                    return False

                ffmpeg_path = shutil.which("ffmpeg")
                if ffmpeg_path is None:
                    self.logger.system("WARN: screen-recording disabled because ffmpeg is not installed", job_id)
                    return False

                width, height = self._get_screen_resolution()

                capture = [
                    "-video_size", f"{width}x{height}",
                    "-f", "x11grab",
                    "-i", display,
                ]

                recording_process = subprocess.Popen(
                    [
                        ffmpeg_path,
                        "-y",
                        *capture,
                        "-framerate", "15",
                        "-vf", drawtext,
                        "-vcodec", "libx264",
                        "-preset", "ultrafast",
                        filename,
                    ],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    start_new_session=True,
                )

            time.sleep(0.2)
            
        except Exception as e:
            # treat a screen recording crash, when screen recording should be used, as a critical error
            raise RobotRuntimeFault(message="unable to start screen recording", phase="queue_for_rpa", active_job=active_job, cause=e) from e

        if recording_process.poll() is not None:
            self.logger.system("ffmpeg exited immediately; recording did not start", job_id)
            raise RobotRuntimeFault(message="unable to start screen recording", phase="queue_for_rpa", active_job=active_job, cause=None)
        
        self.recording_process = recording_process
        self.logger.system("recording started", job_id)
        return True
        
    def stop(self, job_id=None) -> None:
        ''' allow global kill of FFMPEG processes since RobotRuntime is designed to run on a dedicated machine '''
        # written by AI

        self.logger.system("stop recording", job_id)

        recording_process = self.recording_process
        self.recording_process = None

        try:
            if recording_process is not None:
                # try first stop only our own process
                if platform.system() == "Windows":
                    try:
                        recording_process.send_signal(
                            getattr(signal, "CTRL_BREAK_EVENT", signal.SIGTERM)
                        )
                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        subprocess.run(
                            ["taskkill", "/PID", str(recording_process.pid), "/T", "/F"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

                else:
                    # try first stop only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGINT) #type: ignore

                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGKILL) # type: ignore
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

            else:
                # fallback if process object is lost
                if platform.system() == "Windows":
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
                else:
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
        except Exception as err:
            self.logger.system(f"WARN from stop(): {err}", job_id)

    def try_upload_recording(self, job_id, max_attempts=3) -> None:
        ''' upload to a shared drive'''
    
        local_file = f"{self.recordings_in_progress_folder}/{job_id}.mp4"
        local_file = Path(local_file)

        if not local_file.exists():
            self.logger.system(f"no recording file found to upload", job_id)
            return
        
        remote_path = Path(self.recordings_destination_folder) / f"{job_id}.mp4"
        remote_path.parent.mkdir(parents=True, exist_ok=True)

        for attempt in range(max_attempts):
            try:
                shutil.copy2(local_file, remote_path)
                self.logger.system(f"upload successful: {remote_path}", job_id)
                try: os.remove(local_file)
                except Exception as err: self.logger.system(f"{err}")

                return

            except Exception as e:
                self.logger.system(f"upload attempt {attempt+1}/{max_attempts} failed: {e}", job_id)
                time.sleep(attempt + 1)
        
        self.logger.system(f"upload failed: {remote_path}", job_id)

    def cleanup_aborted_recordings(self):
        """Upload or clean up recordings left behind by aborted runs."""

        directory = Path(self.recordings_in_progress_folder)
        if not directory.exists():
            return
        
        for file in directory.iterdir():

            if file.is_file() and file.suffix == ".mp4":
                job_id = file.stem
                self.logger.system(f"cleanup upload started", job_id)
                self.try_upload_recording(job_id)

    def _get_screen_resolution(self):
        # written by AI
        try:
            output = subprocess.check_output(["xrandr"], text=True)
            for line in output.splitlines():
                if "*" in line:
                    res = line.split()[0]
                    return res.split("x")
        except Exception:
            pass

        # fallback: Tkinter
        try:
            root = tk.Tk()
            root.withdraw()
            width = root.winfo_screenwidth()
            height = root.winfo_screenheight()
            root.destroy()
            return str(width), str(height)
        except Exception:
            pass

        return "1920", "1080"


class FriendsRepository:
    '''Access-control source for personal_inbox'''

    def __init__(self, friends_path, organisation_domain, allowed_job_names) -> None:
        self.friends_path = friends_path
        self.organisation_domain = organisation_domain
        self.friends_filename = Path(friends_path).name
        self.allowed_job_names = allowed_job_names
        self.access_by_email: dict[str, set[str]] = {}
        self.access_file_mtime: float | None = None

    def reload_if_modified(self) -> bool:
        '''Reload access file if changed.'''
        # written by AI

        self._ensure_friends_file_exists()

        mtime = os.path.getmtime(self.friends_path)
        if self.access_file_mtime == mtime:
            return False

        new_access = self._load_access_file()
        self._validate_friends_access(new_access)

        self.access_by_email = new_access
        self.access_file_mtime = mtime

        return True

    def is_allowed_sender(self, email_address: str | None) -> bool:

        if not email_address:
            return False
        
        email = email_address.strip().lower()        
        return email in self.access_by_email

    def has_job_access(self, email_address: str, job_name: str) -> bool:
        email = email_address.strip().lower()
        job = job_name.strip().lower()
        return job in self.access_by_email.get(email, set())

    def _ensure_friends_file_exists(self) -> None:
        '''Create a template access file if missing.'''
        if os.path.exists(self.friends_path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "email"
        ws["B1"] = "ping"
        ws["C1"] = "qty_adjust"

        ws["A2"] = "alice@example.com"
        ws["B2"] = "x"

        ws["A3"] = "bob@example.com"
        ws["B3"] = "x"
        ws["C3"] = "x"

        wb.save(self.friends_path)
        wb.close()

    def _load_access_file(self) -> dict[str, set[str]]:
        '''
        Reads access file and returns eg:

        {
            "alice@example.com": {"ping"},
            "bob@example.com": {"ping", "qty_adjust"}
        }
        '''
        # written by AI

        wb = load_workbook(self.friends_path, data_only=True)
        try:
            ws = wb.active
            assert ws is not None

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise ValueError(f"{self.friends_filename} contains no users")

            header = rows[0]
            self._validate_friends_header(header)
            access_map: dict[str, set[str]] = {}

            for row in rows[1:]:
                email_cell = row[0]
                if email_cell is None:
                    continue

                email = str(email_cell).strip().lower()
                if not email:
                    continue

                permissions: set[str] = set()

                for col in range(1, len(header)):
                    jobname = header[col]
                    if jobname is None:
                        continue

                    jobname = str(jobname).strip().lower()
                    cell = row[col] if col < len(row) else None

                    if cell is None:
                        continue

                    if str(cell).strip().lower() == "x":
                        permissions.add(jobname)

                access_map[email] = permissions

            return access_map
        finally:
            wb.close()

    def _validate_friends_access(self, access_map: dict[str, set[str]]) -> None:
        if not isinstance(access_map, dict):
            raise ValueError("access_map must be dict")

        valid_job_names = self.allowed_job_names

        for email, permissions in access_map.items():
            if not isinstance(email, str):
                raise ValueError(f"invalid email key type: {email}")

            email_normalized = email.strip().lower()
            if not email_normalized:
                raise ValueError("empty email in access_map")

            if "@" not in email_normalized:
                raise ValueError(f"invalid email in {self.friends_filename}: {email}")

            if not isinstance(permissions, set):
                raise ValueError(f"permissions must be set for {email}")

            if not email_normalized.endswith(self.organisation_domain):
                raise ValueError(f"{email} is outside domain {self.organisation_domain} and therefore not allowed in {self.friends_filename}.")

            invalid_permissions = permissions - valid_job_names
            if invalid_permissions:
                print(
                    f"WARN! {email} in {self.friends_filename} has access to {sorted(invalid_permissions)}, but this/these job type(s) are not found"
                    f" (activated jobs: {sorted(valid_job_names)})"
                )
            
    def _validate_friends_header(self, header_row) -> None:
        if not header_row or str(header_row[0]).strip().lower() != "email":
            raise ValueError(f"{self.friends_filename} column A must be 'email'")

        '''
        valid_job_names = self.allowed_job_names

        for col in range(1, len(header_row)):
            jobname = header_row[col]
            if jobname is None:
                continue

            jobname_str = str(jobname).strip().lower()
            if jobname_str not in valid_job_names:
                print(
                    f"WARN! job type {jobname_str} in {self.friends_filename} is not found and access will not be considered (activated jobs: {sorted(valid_job_names)})"
                )
        '''


class NetworkService:
    """Check whether the machine currently has access to the required company network resources."""

    def __init__(self, logger, network_healthcheck_path) -> None:
        self.logger = logger
        self.network_state = False
        self.next_network_check_time = 0
        self.network_healthcheck_path = network_healthcheck_path

    def has_network_access(self) -> bool:

        now = time.time()

        if now < self.next_network_check_time:
            return self.network_state

        try:
            if self.network_healthcheck_path is None: # demo assumption
                online = True                         # demo assumption
            else:
                os.listdir(self.network_healthcheck_path)
                online = True 

        except Exception:
            online = False
            
        if online != self.network_state:
            self.network_state = online

            if online:
                self.logger.system("network restored")
            else:
                self.logger.system(f"WARN: network lost")

        # check once every minute if offline, once every 10 min if online
        if online:
            self.next_network_check_time = now + 600
        else:
            self.next_network_check_time = now + 60
        
        return online


class AuditRepository:
    ''' handles an audit-style activity log '''

    def __init__(self, logger, audit_db_path) -> None:
        self.logger = logger
        self.audit_db_path = audit_db_path

    def ensure_db_exists(self) -> None:
        
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
           
            cur.execute('''
                CREATE TABLE IF NOT EXISTS audit_log
                         (
                        job_id INTEGER PRIMARY KEY, 
                        job_name TEXT, 
                        lifecycle_status TEXT, 
                        email_address TEXT, 
                        email_subject TEXT, 
                        source_ref TEXT,
                        started_at_date TEXT, 
                        started_at_time TEXT, 
                        updated_at_time TEXT, 
                        final_reply_sent INTEGER NOT NULL DEFAULT 0,
                        source_type TEXT,
                        error_code TEXT, 
                        error_message TEXT,
                        rpatool_payload TEXT,
                        request_summary TEXT
                        )
                        ''')

    def insert(self, active_job: ActiveJob, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None=None, final_reply_sent=None, error_code=None, error_message=None,) -> None:
        # use for new row

        fields = self._build_audit_fields(
            job_id=active_job.job_id,
            email_address=active_job.email_address,
            email_subject=active_job.email_subject,
            source_ref=active_job.source_ref,
            job_name=active_job.job_name,
            started_at_date=started_at_date,
            started_at_time=started_at_time,
            lifecycle_status=lifecycle_status,
            final_reply_sent=final_reply_sent,
            source_type=active_job.source_type,
            error_code=error_code,
            error_message=error_message,
            rpatool_payload=json.dumps(active_job.rpatool_payload) if active_job.rpatool_payload is not None else None,
            request_summary=active_job.request_summary,
        )
        
        columns = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"INSERT INTO audit_log ({columns}) VALUES ({placeholders})",
                tuple(fields.values())
            )

    def final_reply_sent(self, job_id: int) -> bool:
        existing = self.get_row_by_id(job_id)
        if existing and existing.get("final_reply_sent") == 1:
            return True
        return False

    def parse_from_jobaudit(self, job_id: int) -> ActiveJob:
        row = self.get_row_by_id(job_id)

        if row is None:
            raise ValueError(f"job_id {job_id} not found")

        source_ref = row.get("source_ref")
        source_type = row.get("source_type")
        job_name = row.get("job_name")
        rpatool_payload_raw = row.get("rpatool_payload")

        if not source_ref:
            raise ValueError(f"job_id {job_id} missing source_ref")

        if not source_type:
            raise ValueError(f"job_id {job_id} missing source_type")

        if not job_name:
            raise ValueError(f"job_id {job_id} missing job_name")


        return ActiveJob(
            job_id=job_id,
            job_name=job_name,
            source_ref=source_ref,
            source_type=source_type,
            email_address=row.get("email_address"),
            email_subject=row.get("email_subject"),
            email_body="[ORIGINAL MESSAGE NOT STORED]",
            rpatool_payload= json.loads(rpatool_payload_raw) if rpatool_payload_raw else None,
            request_summary=row.get("request_summary"),
        )

    def count_done_jobs_today(self) -> int:
        today = datetime.date.today().isoformat()

        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ?
                AND lifecycle_status = 'DONE'
            ''', (today,))
            
            result = cur.fetchone()[0]

        return result

    def has_been_processed_today(self, source_ref) -> bool:

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ? AND source_ref = ?
                ''',
                (today, source_ref,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0

    def has_sender_job_today(self, active_job: ActiveJob) -> bool:

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE started_at_date = ? AND email_address = ? AND job_id != ?
                ''',
                (today, active_job.email_address, active_job.job_id,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0

    def get_row_by_id(self, job_id: int) -> dict | None:
        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                "SELECT * FROM audit_log WHERE job_id = ?",
                (job_id,),
            )
            row = cur.fetchone()

        return dict(row) if row is not None else None

    def get_latest_row_by_source_ref(self, source_ref: str) -> dict | None:
        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                "SELECT * FROM audit_log WHERE source_ref = ? ORDER BY job_id DESC LIMIT 1",
                (source_ref,),
            )
            row = cur.fetchone()

        return dict(row) if row is not None else None
    
    def get_latest_job_id(self) -> int:
        with self._connect_with_retry() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id
                FROM audit_log
                ORDER BY job_id DESC
                LIMIT 1
            ''')
            row = cur.fetchone()

        return row[0] if row is not None else 0

    def get_personal_pending_reply_jobs(self) -> list[dict]:
        source_type: SourceType = "personal_inbox"

        with self._connect_with_retry() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT job_id, source_type, email_address, email_subject, source_ref, lifecycle_status, error_code, error_message
                FROM audit_log
                WHERE source_type = ?
                AND COALESCE(final_reply_sent, 0) = 0
                ORDER BY job_id
                ''',
                (source_type,)
            )
            rows = cur.fetchall()

        list_of_dicts = [dict(row) for row in rows]

        return list_of_dicts

    def mark_running(self, job_id):
        lifecycle_status: LifecycleStatus = "RUNNING"
        self._update(
            job_id=job_id,
            lifecycle_status=lifecycle_status
            )
        self.logger.system("marked RUNNING", job_id)
        
    def mark_verifying(self, job_id):
        lifecycle_status: LifecycleStatus = "VERIFYING"
        self._update(
            job_id=job_id, 
            lifecycle_status=lifecycle_status,
            )
        self.logger.system("marked VERIFYING", job_id)

    def mark_done(self, job_id):
        self._update(
            job_id=job_id, 
            lifecycle_status="DONE", 
            )
        self.logger.system("marked DONE", job_id)
    
    def mark_failed(self, job_id, error_code: LifecycleErrorCode, error_message):
        lifecycle_status: LifecycleStatus = "FAIL"
        self._update(
            job_id=job_id, 
            lifecycle_status=lifecycle_status, 
            error_code=error_code, 
            error_message=error_message, 
            )
        self.logger.system("marked FAIL", job_id)
        
    def mark_final_reply_sent(self, job_id):
        self._update(
            job_id=job_id, 
            final_reply_sent=True 
            )

    def _connect_with_retry(self) -> sqlite3.Connection:
  
        max_retries = 3
        for attempt in range(max_retries):
            try:
                conn = sqlite3.connect(self.audit_db_path, timeout=10)
                return conn
            except sqlite3.OperationalError as e:
                self.logger.system(f"WARN: {e}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(0.5)


        return sqlite3.connect("unreachable")  # to satisfy pylance

    def _build_audit_fields(self, job_id, email_address=None, email_subject=None, source_ref=None, job_name: JobName | None = None, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None = None, final_reply_sent=None, source_type: SourceType | None = None, error_code=None, error_message=None, rpatool_payload=None, request_summary=None) -> dict:
        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_name": job_name,
            "started_at_date": started_at_date,
            "started_at_time": started_at_time,
            "updated_at_time": datetime.datetime.now().strftime("%H:%M:%S"),
            "lifecycle_status": lifecycle_status,
            "final_reply_sent": final_reply_sent,
            "source_type": source_type,
            "error_code": error_code,
            "error_message": error_message,
            "rpatool_payload": rpatool_payload,
            "request_summary": request_summary,
        }

        # drop None:s
        fields = {k: v for k, v in all_fields.items() if v is not None}

        gdpr_safe_fields = dict(fields)
        
        if gdpr_safe_fields.get("email_address") is not None: gdpr_safe_fields["email_address"] = "***"
        if gdpr_safe_fields.get("email_subject") is not None: gdpr_safe_fields["email_subject"] = "***"

        suffix = " (GDPR-sanitized)" if fields != gdpr_safe_fields else ""
        self.logger.system(f"received audit fields {gdpr_safe_fields}{suffix}", job_id)

        return fields

    def _update(self, job_id, email_address=None, email_subject=None, source_ref=None, job_name: JobName | None=None, started_at_date=None, started_at_time=None, lifecycle_status: LifecycleStatus | None=None, final_reply_sent=None, source_type:SourceType | None=None, error_code=None, error_message=None, rpatool_payload=None, request_summary=None) -> None:
        # use eg: self.audit.update(job_id=20260311124501, job_name="qty_adjust")
        # intentionally not using active_job

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_name=job_name,
            started_at_date=started_at_date,
            started_at_time=started_at_time,
            lifecycle_status=lifecycle_status,
            final_reply_sent=final_reply_sent,
            source_type=source_type,
            error_code=error_code,
            error_message=error_message,
            rpatool_payload=json.dumps(rpatool_payload) if rpatool_payload is not None else None,
            request_summary=request_summary,
            )
        
        fields.pop("job_id", None)

        if not fields:
            return

        set_clause = ", ".join(f"{k}=?" for k in fields)

        with self._connect_with_retry() as conn:
            cur = conn.cursor()

            cur.execute(
                f"UPDATE audit_log SET {set_clause} WHERE job_id=?",
                (*fields.values(), job_id)
            )

            if cur.rowcount == 0:
                raise ValueError(f"update(): no row in DB with job_id={job_id}")


class LoggerService:
    """ logging functions """

    def __init__(self, dashboard_ui, system_log_path) -> None:
        self.dashboard_ui = dashboard_ui
        self.system_log_path = system_log_path

    def ui(self, text:str, blank_line_before: bool = False) -> None:
        
        try:
            self.dashboard_ui.post_log_line(text, blank_line_before)
        except Exception as err:
            self.system(f"logger.ui fail={err}")

    def system(self, event_text, job_id: int | None=None,):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        event_text = str(event_text)

        # get caller function name
        try:
            frame = sys._getframe(1)
            caller_name = frame.f_code.co_name
            instance = frame.f_locals.get("self")
            if instance is not None:
                class_name = instance.__class__.__name__
                caller = f"{class_name}.{caller_name}()"
            else:
                caller = f"{caller_name}()"

        except Exception:
            caller = "unknown_caller()"
      
        log_line = f"{timestamp} | RR  | job_id={job_id or ''} | {caller} | {event_text}"

        # normalize to single-line
        log_line = " ".join(str(log_line).split())

        last_err = None
        for i in range(7):
            try:
                with open(self.system_log_path, "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
                    f.flush()
                return

            except Exception as err:
                last_err = err
                print(f"WARN: retry {i+1}/7 from log_system():", err)
                time.sleep(i + 1)

        # fallback to print() when log fails        
        print(f"[print fallback] {job_id} {event_text} | {last_err}")  
 

class MailRecoveryService:
    '''stuck/crashed mail handler used in while in safestop'''

    def __init__(self, logger, personal_mailbox, shared_mailbox, audit, friends_repo, notifications, generate_job_id) -> None:
        self.logger = logger
        self.personal_mailbox = personal_mailbox
        self.shared_mailbox = shared_mailbox
        self.audit = audit
        self.job_audit_filename = Path(self.audit.audit_db_path).name
        self.friends_repo = friends_repo
        self.notifications = notifications
        self.friends_filename = Path(self.friends_repo.friends_path).name
        self.generate_job_id = generate_job_id
   
    def recover_faulted_personal_job(self, fault: RobotRuntimeFault, recovery_context: str,) -> None:
        active_job = fault.active_job

        if active_job is None:
            return

        if active_job.source_type != "personal_inbox":
            return
        
        lifecycle_status: LifecycleStatus = "FAIL"
        error_code: LifecycleErrorCode = fault.error_code
        public_error_message: str | None = fault.error_message
        fallback_status: Literal["DONE", "FAIL"]

        
        if active_job.job_id is None:
            active_job.job_id = self.generate_job_id()
        else:
            existing = self.audit.get_row_by_id(active_job.job_id)

            #if self.audit.final_reply_sent(active_job.job_id):
            if existing and existing.get("final_reply_sent"):
                self.logger.system("recovery skipped final reply; already sent", active_job.job_id)
                try:
                    fallback_status = "DONE" if existing.get("lifecycle_status") == "DONE" else "FAIL"
                    self.personal_mailbox.delete(active_job, fallback_status=fallback_status)
                except Exception as err: self.logger.system(err, active_job.job_id)
                return
            

        if not active_job.job_name:
            active_job.job_name = "unknown"

        # Job audit
        try:
            if existing:

                existing_status = existing.get("lifecycle_status")
                if existing_status in {"DONE", "FAIL", "REJECTED"}:
                    # keep original status
                    lifecycle_status = existing_status
                    error_code = existing.get("error_code")
                    public_error_message = existing.get("error_message")
                else:
                    self.audit.mark_failed(
                        job_id=active_job.job_id,
                        error_code=fault.error_code,
                        error_message=fault.error_message or str(fault.cause) or "runtime fault",
                    )
            else:
                self.audit.insert(
                    active_job=active_job,
                    started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
                    started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
                    lifecycle_status=lifecycle_status,
                    error_code=fault.error_code,
                    error_message=fault.error_message or str(fault.cause) or "runtime fault",
                    final_reply_sent=False,
                )
        except Exception as err:
            self.logger.system(err, active_job.job_id)

        # User reply
        try:
            if not self.personal_mailbox.sent_reply_exists(active_job.source_ref):
                self.notifications.send_final_reply(
                    active_job=active_job,
                    lifecycle_status=lifecycle_status,
                    error_code=error_code,
                    public_error_message=public_error_message,
                    recovery_context=recovery_context,
                )

            self.audit.mark_final_reply_sent(active_job.job_id)
            fallback_status = "DONE" if lifecycle_status == "DONE" else "FAIL"
            self.personal_mailbox.delete(active_job, fallback_status=fallback_status)
        except Exception as err:
            self.logger.system(err, active_job.job_id)

    def process_one_personal_mail_in_safestop(self, log_as_recovered=False) -> None:
        """Process one personal inbox email while in safestop."""
        active_job: ActiveJob

        context_text = " recovered" if log_as_recovered else ""

        # process one personal inbox email in degraded mode
        paths = self.personal_mailbox.list_inbox_mail_paths()
        if not paths:
            return
        
        for inbox_path in paths:
            active_job = self.personal_mailbox.parse_mail_file(inbox_path)

            if self.personal_mailbox._has_status_prefix(active_job, "DONE"):
                continue
            if self.personal_mailbox._has_status_prefix(active_job, "FAIL"):
                continue

            self.logger.system(f"found{context_text} source_ref={active_job.source_ref}")
            self.logger.ui(f"email{context_text} from {active_job.email_address}", blank_line_before=True)

            # silent delete non friends
            if not self.friends_repo.is_allowed_sender(active_job.email_address):
                self.logger.ui(f"--> rejected (not in {self.friends_filename})")
                self.personal_mailbox.delete(active_job, fallback_status="DONE")
                self.logger.system(f"silently deleted source_ref={active_job.source_ref} (not in {self.friends_filename})")
                return
            
            # check for email commands
            if self._check_for_stop_command(active_job):
                return
            if self._check_for_restart_command(active_job):
                return

            # reply, audit-log and delete for friends
            active_job.job_id = self.generate_job_id()
            final_reply_sent = False

            try:
                # TODO: add argument active_job.source_ref to final_reply_sent()
                if self.personal_mailbox.sent_reply_exists(active_job.source_ref): # or self.audit.final_reply_sent(job_id):
                    self.logger.system("recovery reply skipped, a final reply to user already exists", active_job.job_id)
                else:
                    self.notifications.send_out_of_service_reply(active_job)
                    final_reply_sent = True
            except Exception as e:
                self.logger.system(e, active_job.job_id)

            if not final_reply_sent:
                self.logger.ui("--> retry later (reply failed)")
                self.logger.system(f"out-of-service reply failed, keeping mail to retry later", active_job.job_id,)
                return

            try: self.insert_recovery_audit_row(active_job, final_reply_sent, recovery_reason="SAFESTOP")
            except Exception as e: self.logger.system(f"job audit error={e}")
            self.logger.system(f"out-of-service reply sent and job audit updated for source_ref={active_job.source_ref}")
            
            try: self.personal_mailbox.delete(active_job, fallback_status="FAIL")
            except Exception as e: self.logger.system(f"WARN: {active_job.source_ref} crashed in personal_mailbox.delete() with error {e}", active_job.job_id,)
            self.logger.ui("--> rejected (safestop)")
        
    def recover_stuck_shared_mail(self, fault: RobotRuntimeFault | None = None) -> None:
        # written by AI
        active_job: ActiveJob | None = None
        job_id: int | None = None
        if fault:
            job_id = fault.handover_file.job_id if fault.handover_file else fault.active_job.job_id if fault.active_job else None

        # 1. Try audit-based recovery first
        if job_id is not None:
            try:
                active_job = self.audit.parse_from_jobaudit(job_id)
            except Exception as e:
                self.logger.system(f"shared recovery: could not parse audit job: {e}", job_id)
                active_job = None
       
       
        if active_job is not None:
            if active_job.source_type != "shared_inbox":
                return

            try:
                self.audit.mark_failed(
                    job_id=active_job.job_id,
                    error_code=fault.error_code if fault else "PRE_HANDOVER_CRASH",
                    error_message=fault.error_message if fault else "unknown shared mail recovery",
                )
            except Exception as e:
                self.logger.system(
                    f"shared audit update failed for faulted job: {e}",
                    active_job.job_id,
                )
            
            self.logger.ui("--> fail (safestop)")
            try:
                self.shared_mailbox.mark_failed(active_job)
                return
            except Exception as e:
                self.logger.system(
                    f"shared mail recovery failed, error={e}",
                    active_job.job_id,
                )


        # 2. Fallback: no usable audit row. Scan inbox for PROCESSING mail.
        for path in self.shared_mailbox.list_inbox_mail_paths():
            
            try:
                active_job_from_scan = self.shared_mailbox.parse_mail_file(path)
            except Exception as e:
                self.logger.system(f"shared recovery: could not parse mail {path}: {e}", job_id)
                continue

            if not self.shared_mailbox._has_status_prefix(active_job_from_scan, "PROCESSING"):
                continue
            
            row = self.audit.get_latest_row_by_source_ref(active_job_from_scan.source_ref)
            if row and row.get("lifecycle_status") != "FAIL":
                    self.audit.mark_failed(
                        job_id=row.get("job_id"),
                        error_code="PRE_HANDOVER_CRASH",
                        error_message="stale shared PROCESSING mail found",
                    )

            active_job_from_scan.job_id = job_id or self.generate_job_id()

            self.logger.ui("--> fail  (safestop)")
            try:
                self.shared_mailbox.mark_failed(active_job_from_scan)
                self.logger.system(
                    f"shared recovery fallback marked FAIL: {active_job_from_scan.source_ref}",
                    active_job_from_scan.job_id,
                )
            except Exception as e:
                self.logger.system(
                    f"shared recovery fallback failed: {e}",
                    active_job_from_scan.job_id,
                )

            return
    
    def insert_recovery_audit_row(self, active_job:ActiveJob, final_reply_sent: bool, recovery_reason,):
        
        if recovery_reason == "SAFESTOP":
            lifecycle_status="REJECTED"
            error_code="IN_SAFESTOP"
            error_message="not accepting new jobs in safestop"
        
        elif recovery_reason == "RECOVERY":
            lifecycle_status="FAIL"
            error_code="PRE_HANDOVER_CRASH"
            error_message="unknown, mail stuck with PROCESSING subject prefix"
        
        else:
            raise ValueError(f"unknown reason: {recovery_reason}")
   
        self.audit.insert(
            active_job=active_job,
            started_at_date=datetime.datetime.now().strftime("%Y-%m-%d"),
            started_at_time=datetime.datetime.now().strftime("%H:%M:%S"),
            lifecycle_status=lifecycle_status,
            error_code=error_code,
            error_message=error_message,
            final_reply_sent = final_reply_sent,
        )

    def _check_for_stop_command(self, active_job: ActiveJob) -> bool:

        if "stop1234" in str(active_job.email_subject).strip().lower():
            self.logger.system(f"stop command received from {active_job.email_address}")
            Path("stop.flag").write_text("", encoding="utf-8")
            
            try: self.notifications.send_command_reply(active_job)
            except Exception as err: self.logger.system(f"{err}")

            try: self.personal_mailbox.delete(active_job, fallback_status="DONE")
            except Exception as err: self.logger.system(f"{err}")

            return True
        
        return False

    def _check_for_restart_command(self, active_job: ActiveJob) -> bool:

        if "restart1234" in str(active_job.email_subject).strip().lower():
            self.logger.system(f"restart command received from {active_job.email_address}")
            Path("restart.flag").write_text("", encoding="utf-8")
            
            try: self.notifications.send_command_reply(active_job)
            except Exception as err: self.logger.system(f"{err}")

            try: self.personal_mailbox.delete(active_job, fallback_status="DONE")
            except Exception as err: self.logger.system(f"{err}")

            return True
        
        return False


class SafestopController:
    """Handle degraded mode, crash recovery, and operator restart/stop commands."""

    def __init__(self, logger, recording, hide_recording_overlay, post_status_update, set_ui_shutdown, check_for_stop_flag, handover_file, mail_recovery, notifications) -> None:
        self.logger = logger
        self.recording = recording
        self._hide_recording_overlay = hide_recording_overlay
        self._post_status_update = post_status_update
        self._set_ui_shutdown = set_ui_shutdown
        self._check_for_stop_flag = check_for_stop_flag
        self.HANDOVER_FILE = handover_file
        self.mail_recovery = mail_recovery
        self.notifications = notifications
        self._degraded_mode_entered = False

    def run_degraded_mode(self, fault: RobotRuntimeFault) -> None:
        '''
        Rules:
        * no job intake
        * mail-flow inactivated
        * query-flow inactivated
        * 'safestop' status text in UI
        * STOP and RESTART commands available 
        * REJECTED reply to new emails from users in friends.xlsx
        * changes in friends.xlsx access list will have no effect
        * notification email is sent to rpa admin
        '''
        
        if self._degraded_mode_entered: return
        self._degraded_mode_entered = True

        handover_file = fault.handover_file
        active_job = fault.active_job
        job_id = handover_file.job_id if handover_file else active_job.job_id if active_job else None

        # overwrite job_queued to stop RPA tool from possibly starting the job
        if handover_file is not None and handover_file.state == "job_queued":
            try:
                handover_file.state="safestop"
                self._write_handover_directly(handover_file)

            except Exception:
                try: os.remove(self.HANDOVER_FILE)
                except Exception as e: self.logger.system(e)

        crash_report = (
                "ROBOTRUNTIME CRASHED\n\n"
                f"phase={fault.phase}\n"
                f"error_code={fault.error_code}\n"
                f"fault={fault.error_message}\n\n"
                f"{fault.traceback_text}"
            )

        if active_job is not None:
            crash_report += (
                f"\n\n...while handling active_job:"
                f"\nsource_type={active_job.source_type}"
                f"\njob_name={active_job.job_name}"
                f"\nsource_ref={active_job.source_ref}"
            )
        
        if handover_file is not None:
            crash_report += (
                f"\n\n...while working on job_name={handover_file.job_name} "
                f"with rpatool_payload=\n{handover_file.rpatool_payload}"
            )

        self.logger.system(crash_report, job_id)
        self.logger.ui(f"--> CRASH! All automations are stopped. Robot admin is notified")

        if active_job is not None:
            crash_report += (
                f"\n\nExtra admin info omitted from system log:"
                f"\nemail_address={active_job.email_address}"
                f"\nemail_subject={active_job.email_subject}"
            )

        try: self.notifications.send_admin_alert(crash_report, critical=True)
        except Exception as e: self.logger.system(e, job_id)

        try: self.recording.stop(job_id)
        except Exception as e: self.logger.system(e, job_id)
        
        try: self._hide_recording_overlay()
        except Exception as e: self.logger.system(e, job_id)

        try: self.recording.cleanup_aborted_recordings()
        except Exception as e: self.logger.system(e, job_id)

        try: self.mail_recovery.recover_faulted_personal_job(fault, recovery_context="safestop")
        except Exception as e: self.logger.system(e, job_id)

        try: self.mail_recovery.recover_stuck_shared_mail(fault)
        except Exception as e: self.logger.system(e, job_id)

        # placeholder for recovery logic for post_handover crash/mismatch for query jobs

        try: self._post_status_update("safestop")
        except Exception as e: self.logger.system(e, job_id)
    
        
        self._enter_degraded_loop()
    
    def _check_for_restart_flag(self,) -> None:
        restartflag = "restart.flag"

        if os.path.isfile(restartflag):
            try: os.remove(restartflag)
            except Exception as err: self.logger.system(f"{err}")

            self.logger.system(f"restart-command received from {restartflag}")
            
            try:
                self._write_handover_directly(HandoverFile(state="idle"))
            except Exception as e:
                self.logger.system(f"could not reset handover before restart: {e}")
                os._exit(1)
            
            self._restart_application()

    def _write_handover_directly(self, handover_file: HandoverFile) -> None:
        '''write w/o using full handover in degraded mode'''
        handover_data = asdict(handover_file)

        temp_path = f"{self.HANDOVER_FILE}.tmp"

        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(handover_data, f, indent=2)
            f.flush()
            os.fsync(f.fileno())

        os.replace(temp_path, self.HANDOVER_FILE)

    def _enter_degraded_loop(self) -> Never:
        ''' follow policy to always reply to known users'''  

        self.logger.system("running")
        log_as_recovered = True

        while True:
            try:
                time.sleep(1)
                self._check_for_stop_flag()
                self._check_for_restart_flag()
                self.mail_recovery.process_one_personal_mail_in_safestop(log_as_recovered)
                
            except Exception as e:
                self.logger.system(e)
            
            finally:
                log_as_recovered = False

    def _restart_application(self) -> Never:
        ''' written by AI '''
        self.logger.system("restarting application in new visible terminal")

        try:
            self._set_ui_shutdown()
        except Exception:
            pass

        try:
            script_path = os.path.abspath(sys.argv[0])

            if platform.system() == "Windows":
                subprocess.Popen(
                    [sys.executable, script_path],
                    creationflags=subprocess.CREATE_NEW_CONSOLE # type: ignore
                )

            else:
                python_cmd = f'"{sys.executable}" "{script_path}"'

                terminal_active_jobs = [
                    ["gnome-terminal", "--", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xfce4-terminal", "--hold", "-e", python_cmd],
                    ["konsole", "-e", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xterm", "-hold", "-e", python_cmd],
                ]

                launched = False
                for cmd in terminal_active_jobs:
                    try:
                        subprocess.Popen(cmd)
                        launched = True
                        break
                    except FileNotFoundError:
                        continue

                if not launched:
                    raise RuntimeError("No supported terminal emulator found for restart")

        except Exception as e:
            self.logger.system(e)
            os._exit(1)

        time.sleep(1)
        os._exit(0)


# ============================================================
# UI
# ============================================================

class DashboardUI:
    """Tkinter dashboard for visibility of runtime status and logs"""

    # colors
    BG = "#000000"
    TEXT = "#F5F5F5"
    MUTED = "#A0A0A0"
    GREEN = "#22C55E"
    GREEN_2 = "#16A34A"
    GREEN_3 = "#15803D"
    RED = "#DC2626"
    YELLOW = "#FACC15"
    SCROLL_TROUGH = "#0F172A"
    SCROLL_BG = "#1E293B"
    SCROLL_ACTIVE = "#475569"

    # fonts
    FONT_STATUS = ("Arial", 70, "bold")
    FONT_COUNTER = ("Segoe UI", 100, "bold")
    FONT_SMALL = ("Arial", 14, "bold")
    FONT_LOG = ("DejaVu Sans Mono", 20)
    FONT_RECORDING = ("Arial", 20, "bold")

    # sizes
    ROOT_PADX = 50
    SCROLLBAR_WIDTH = 23

    RECORDING_WIDTH = 250
    RECORDING_HEIGHT = 110
    RECORDING_MARGIN_RIGHT = 30


    def __init__(self):
        self._build_root(self.BG)
        self._build_header(self.BG, self.TEXT)
        self._build_body(self.BG, self.TEXT)
        self._build_footer(self.BG, self.TEXT)

        #self._debug_grid(self.root)

    def run(self) -> None:
        self.root.mainloop()

    def post_status_update(self, status: DashboardStatus) -> None:
        self.root.after(0, lambda: self._apply_status_update(status))

    def post_log_line(self, text: str, blank_line_before: bool = False) -> None:
        self.root.after(0, lambda: self._append_ui_log(text, blank_line_before))

    def post_show_recording_overlay(self) -> None:
        self.root.after(0, self._show_recording_overlay)

    def post_hide_recording_overlay(self) -> None:
        self.root.after(0, self._hide_recording_overlay)

    def post_jobs_done_today(self, n: int) -> None:
        self.root.after(0, lambda: self._apply_jobs_done_today(n))

    def post_shutdown(self, delay=0) -> None:
        self.root.after(delay, self._shutdown)

    def _debug_grid(self, widget):
        ''' highlights all grids with red '''
        for child in widget.winfo_children():
            try: child.configure(highlightbackground="red", highlightthickness=1)
            except Exception: pass
            self._debug_grid(child)

    def _build_root(self, bg_color):
        self.root = tk.Tk()

        w = self.root.winfo_screenwidth()
        h = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+0+0")

        self.root.configure(bg=bg_color, padx=self.ROOT_PADX)
        self._closing = False
        self.root.protocol("WM_DELETE_WINDOW", self._on_close_attempt)

        self.root.title('RPA dashboard')
        self._create_recording_overlay()

        # layout using grid
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

    def _build_header(self, bg_color, text_color):
        self.header = tk.Frame(self.root, bg=bg_color)

        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_columnconfigure(2, weight=1)
        self.header.grid_rowconfigure(0, weight=1)

        # Header content
        self.rpa_text_label = tk.Label(
            self.header,
            text="RPA:",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_STATUS,
        )  
        self.rpa_text_label.grid(row=0, column=0, padx=16, pady=16, sticky="w")

        self.rpa_status_label = tk.Label(
            self.header,
            text="",
            fg=self.RED,
            bg=bg_color,
            font=self.FONT_STATUS,
        )
        self.rpa_status_label.grid(row=0, column=1, padx=16, pady=16, sticky="w")

        self.status_dot = tk.Label(
            self.header,
            text="",
            fg=self.GREEN,
            bg=bg_color,
            font=("Arial", 50, "bold"),
        )
        self.status_dot.grid(row=0, column=2, sticky="w")

        # jobs done today (counter + label in same grid)
        self.jobs_counter_frame = tk.Frame(self.header, bg=bg_color)
        self.jobs_counter_frame.grid(row=0, column=3, sticky="ne", padx=40, pady=30)
        self.jobs_counter_frame.grid_rowconfigure(0, weight=1)
        self.jobs_counter_frame.grid_columnconfigure(0, weight=1)

        # normal view (jobs done today)
        self.jobs_normal_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_normal_view.grid(row=0, column=0, sticky="nsew")
        self.jobs_normal_view.grid_columnconfigure(0, weight=1)

        self.jobs_done_label = tk.Label(
            self.jobs_normal_view,
            text="0",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_COUNTER,
            anchor="e",
            justify="right",
        )
        self.jobs_done_label.grid(row=0, column=0, sticky="e")

        self.jobs_counter_text = tk.Label(
            self.jobs_normal_view,
            text="jobs done today",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.jobs_counter_text.grid(row=1, column=0, sticky="e", pady=(0, 6))

        # safestop view (big X)
        self.jobs_error_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_error_view.grid(row=0, column=0, sticky="nsew")

        self.safestop_x_label = tk.Label(
            self.jobs_error_view,
            text="X",
            bg=self.RED,
            fg="#FFFFFF",
            font=self.FONT_COUNTER,
        )  # text="✖",
        self.safestop_x_label.pack(expand=True)

        # show normal view at startup
        self.jobs_normal_view.tkraise()

        # 'online'-status animation
        self._online_animation_after_id = None
        self._online_pulse_index = 0

        # 'working...'-status animation
        self._working_animation_after_id = None
        self._working_dots = 0

    def _build_body(self, bg_color, text_color):
        self.body = tk.Frame(self.root, bg=bg_color)
        self.body.grid(row=1, column=0, sticky="nsew")
        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        # body content
        log_and_scroll_container = tk.Frame(self.body, bg=bg_color)
        log_and_scroll_container.grid(row=0, column=0, sticky="nsew")
        log_and_scroll_container.grid_rowconfigure(0, weight=1)
        log_and_scroll_container.grid_columnconfigure(0, weight=1)

        # the right-hand side scrollbar
        scrollbar = tk.Scrollbar(
            log_and_scroll_container,
            width=self.SCROLLBAR_WIDTH,
            troughcolor=self.SCROLL_TROUGH,
            bg=self.SCROLL_BG,
            activebackground=self.SCROLL_ACTIVE,
            bd=0,
            highlightthickness=0,
            relief="flat",
        )
        scrollbar.grid(row=0, column=1, sticky="ns")

        # the 'console'-style log
        self.log_text = tk.Text(
            log_and_scroll_container,
            yscrollcommand=scrollbar.set,
            bg=bg_color,
            fg=text_color,
            insertbackground="black",
            font=self.FONT_LOG,
            wrap="none",
            state="disabled",
            bd=0,
            highlightthickness=0,
        )  # glow highlightbackground="#1F2937", highlightthickness=1
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=self.log_text.yview)

    def _build_footer(self, bg_color, text_color):
        self.footer = tk.Frame(self.root, bg=bg_color)
        self.footer.grid(row=2, column=0, sticky="nsew")
        self.footer.grid_rowconfigure(0, weight=1)
        self.footer.grid_columnconfigure(0, weight=1)

        # footer content
        self.last_activity_label = tk.Label(
            self.footer,
            text="last activity: xx:xx",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.last_activity_label.grid(row=0, column=1, padx=8, pady=16)

    def _apply_status_update(self, status: DashboardStatus | None = None):

        # stops any ongoing animations
        self._stop_online_animation()
        self._stop_working_animation()
        self.status_dot.config(text="")

        # changes text
        if status == "online":
            self.rpa_status_label.config(text="online", fg=self.GREEN)
            self.jobs_normal_view.tkraise()
            self.status_dot.config(text="●")
            self._start_online_animation()

        elif status == "no_network":
            self.rpa_status_label.config(text="no network", fg=self.RED)
            self.jobs_normal_view.tkraise()

        elif status == "working":
            self.rpa_status_label.config(text="working...", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()
            self._start_working_animation()

        elif status == "safestop":
            self.rpa_status_label.config(text="safestop", fg=self.RED)
            self.jobs_error_view.tkraise()

        elif status == "out_of_office":
            self.rpa_status_label.config(text="out of office", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()

    def _apply_jobs_done_today(self, n) -> None:
        self.jobs_done_label.config(text=str(n))

    def _create_recording_overlay(self) -> None:
        # written by AI
        self.recording_win = tk.Toplevel(self.root)
        self.recording_win.withdraw()                # hidden at start
        self.recording_win.overrideredirect(True)    # no title/border
        self.recording_win.configure(bg="black")

        try:
            self.recording_win.attributes("-topmost", True)
        except Exception:
            pass

        width = self.RECORDING_WIDTH
        height = self.RECORDING_HEIGHT
        x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(
            self.recording_win,
            bg="black",
            highlightbackground="#444444",
            highlightthickness=1,
            bd=0,
        )
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(
            frame,
            width=44,
            height=44,
            bg="black",
            highlightthickness=0,
            bd=0,
        )
        canvas.place(x=18, y=33)
        canvas.create_oval(4, 4, 40, 40, fill=self.RED, outline=self.RED)

        label = tk.Label(
            frame,
            text="RECORDING",
            fg="#FFFFFF",
            bg="black",
            font=self.FONT_RECORDING,
            anchor="w",
        )
        label.place(x=75, y=33)

    def _show_recording_overlay(self) -> None:
        # written by AI
        try:
            width = self.RECORDING_WIDTH
            height = self.RECORDING_HEIGHT
            x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

            self.recording_win.deiconify()
            self.recording_win.lift()

            try:
                self.recording_win.attributes("-topmost", True)
            except Exception:
                pass
        except Exception:
            pass

    def _hide_recording_overlay(self) -> None:
        # hides recording window
        try:
            self.recording_win.withdraw()
        except Exception:
            pass

    def _start_working_animation(self):
        if self._working_animation_after_id is None:
            self._animate_working()

    def _animate_working(self):
        # written by AI
        states = ["working", "working.", "working..", "working..."]
        self._working_dots = (self._working_dots + 1) % len(states)
        self.rpa_status_label.config(text=states[self._working_dots])
        self._working_animation_after_id = self.root.after(500, self._animate_working)

    def _stop_working_animation(self):
        if self._working_animation_after_id is not None:
            self.root.after_cancel(self._working_animation_after_id)
            self._working_animation_after_id = None
            self._working_dots = 0

    def _start_online_animation(self):
        if self._online_animation_after_id is None:
            self._online_pulse_index = 0
            self._animate_online()

    def _animate_online(self):
        # green pulse animation
        colors = [self.GREEN, self.GREEN_2, self.BG, self.GREEN_3, self.GREEN_2]
        color = colors[self._online_pulse_index]

        self.status_dot.config(fg=color)

        self._online_pulse_index = (self._online_pulse_index + 1) % len(colors)
        self._online_animation_after_id = self.root.after(1000, self._animate_online)

    def _stop_online_animation(self):
        if self._online_animation_after_id is not None:
            self.root.after_cancel(self._online_animation_after_id)
            self._online_animation_after_id = None

    def _append_ui_log(self, log_line: str, blank_line_before: bool = False) -> None:

        self.log_text.config(state="normal")  # open for edit
        now = datetime.datetime.now().strftime("%H:%M")

        if blank_line_before:
            self.log_text.insert("end", "\n")

        self.log_text.insert("end", f"[{now}] {log_line}\n")

        self.log_text.config(state="disabled")  # closing edit
        self.log_text.see("end")

    def _on_close_attempt(self):
        message="Use STOP-button in the RPA tool next time (or in the rpa tool simulator)"
        self.post_log_line(message, blank_line_before=True)
        print(message)
        self.post_shutdown(delay=2000)

    def _shutdown(self) -> None:
        if self._closing:
            return

        self._closing = True

        self.root.destroy()


# ============================================================
# MAIN ENTRYPOINT
# ============================================================

class RobotRuntime:
    """Main orchestration runtime."""

    def __init__(self, ui, config):
        self.ui = ui
        self.config = config
        
        self.dashboard_status: DashboardStatus | None = None
        self.prev_state: HandoverState | None = None
        self.rpa_tool_claim_started_at: float | None = None
        self.rpa_tool_execution_started_at: float | None = None
        self.next_queryflow_check_time = 0

        self.logger = LoggerService(self.ui, config.system_log_path)
        self.handover = HandoverRepository(self.logger, config.handover_file)
        self.audit = AuditRepository(self.logger, config.audit_db_path)
        self.network_service = NetworkService(self.logger, config.network_healthcheck_path)
        self.recording = RecordingService(self.logger, config.recordings_in_progress_folder, config.recordings_destination_folder)

        backends = {
            "personal_mailbox": DemoMailBackend("personal_inbox"),
            "shared_mailbox": DemoMailBackend("shared_inbox"),
            "erp_backend": DemoErpBackend(),
        }

        # use custom_backends.py to add custom backend(s) (and use same name to override demo, eg. 'erp_backend')
        if build_custom_backends:
            custom_backends = build_custom_backends()
            backends.update(custom_backends)

        self.personal_mailbox = backends["personal_mailbox"]
        self.shared_mailbox = backends["shared_mailbox"]
        self.erp_backend = backends["erp_backend"]

        self.personal_mail_handlers = {"ping": PingHandler(self.logger),}
        self.shared_mail_handlers = {}
        self.query_handlers = {}

        if build_custom_query_handlers is not None:
            self.query_handlers.update(build_custom_query_handlers(self.logger, self.erp_backend,))

        if build_custom_personal_mail_handlers is not None:
            self.personal_mail_handlers.update(build_custom_personal_mail_handlers(self.logger,))

        if build_custom_shared_mail_handlers is not None:
            self.shared_mail_handlers.update(build_custom_shared_mail_handlers(self.logger,))
        
        self.job_handlers = {
            **self.personal_mail_handlers,
            **self.shared_mail_handlers,
            **self.query_handlers,
        }

        self._validate_job_handlers_registry()
        
        self.friends_repo = FriendsRepository(config.friends_path, config.organisation_domain, allowed_job_names=set(self.personal_mail_handlers.keys()))
        self.notifications = UserNotificationService(self.logger, self.personal_mailbox, self.friends_repo, config)
        self.job_lifecycle = JobLifecycleService(self.logger, self.handover, self.ui.post_show_recording_overlay, self.recording, self.audit, self.notifications, self.personal_mailbox, self.shared_mailbox, self.job_handlers, self.ui.post_hide_recording_overlay, self.generate_job_id)
        self.mail_flow = MailFlow(self.logger, self.friends_repo, self.audit, self._is_within_operating_schedule, self.network_service, self.personal_mail_handlers, self.shared_mail_handlers, self.job_lifecycle, self.personal_mailbox, self.shared_mailbox,)
        self.query_flow = QueryFlow(self.logger, self.query_handlers, self.audit, self.job_lifecycle, self._is_within_operating_schedule)
        self.mail_recovery = MailRecoveryService(self.logger, self.personal_mailbox, self.shared_mailbox, self.audit, self.friends_repo, self.notifications, self.generate_job_id) 
        self.safestop_controller = SafestopController(self.logger, self.recording, self.ui.post_hide_recording_overlay, self.ui.post_status_update, self.ui.post_shutdown, self._check_for_stop_flag, config.handover_file, self.mail_recovery, self.notifications) 
       
    def runtime_loop(self) -> None:
        handover_file: HandoverFile | None = None
        phase: RuntimePhase = "startup"

        try:
            self._startup_sequence()
            
            while True:
                self._check_for_stop_flag()

                handover_file = self.handover.read()
                phase = self._map_state_to_phase(handover_file.state)
                self._handle_state_change(handover_file)
                self._enforce_timeouts(handover_file)
                
                # Dispatch
                if handover_file.state == "idle":                 # RobotRuntime owns the workflow
                    self._run_job_intake()

                elif handover_file.state == "job_queued":         # RPA tool owns the workflow
                    pass

                elif handover_file.state == "job_running":        # RPA tool owns the workflow
                    pass

                elif handover_file.state == "job_verifying":      # RobotRuntime owns the workflow
                    self.job_lifecycle.complete_from_handover(handover_file)
                    self._refresh_jobs_done_counter()

                elif handover_file.state == "safestop":           # RobotRuntime owns the workflow
                    raise RobotRuntimeFault(message="unexpected safestop state from RPA tool", phase=phase, error_code="RPA_TOOL_CRASH", handover_file=handover_file,) 

                time.sleep(self.config.poll_interval)


        except RobotRuntimeFault as fault:
            fault.traceback_text = traceback.format_exc()
            self.safestop_controller.run_degraded_mode(fault)

        except Exception as err:
            fault = RobotRuntimeFault(
                message=str(err),
                phase=phase,
                handover_file=handover_file,
                error_code="CODE_ERROR",
                cause=err,
                traceback_text=traceback.format_exc(),
            )
            self.safestop_controller.run_degraded_mode(fault)
      
    def generate_job_id(self) -> int:
        '''Works under single-machine assumption'''

        now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        job_id = int(now)

        try:
            last_job_id = self.audit.get_latest_job_id()
        except Exception as e:
            last_job_id = int(now) + 5
            self.logger.system(f"WARN: using fallback last_job_id due to audit error={e}")

        job_id = max(job_id, last_job_id + 1)

        self.logger.system(f"generated job_id", job_id)
        return job_id

    def _startup_sequence(self):
        handover_file: HandoverFile | None = None

        try:
            self.logger.system(f"RobotRuntime started, version={VERSION}, pid={os.getpid()}")

            # cleanup
            for f in ["stop.flag", "restart.flag"]:
                try: os.remove(f)
                except Exception: pass
       
            # write 'idle' to avoid unnecessary errors in demo mode (the intended way is RPA tool creates handover.json)
            # Cold-start policy
            self.handover.write(HandoverFile(state="idle"))

            handover_file = self.handover.read()
            if handover_file.state != "idle":
                raise ValueError(
                    f"Expected {self.handover.handover_file} to start in idle, got {handover_file.state}"
                )        

            atexit.register(self.recording.stop)
    
            self.audit.ensure_db_exists()
            self.network_service.has_network_access()
            self.recording.stop() # global ffmpeg kill
            self.recording.cleanup_aborted_recordings()
            self.friends_repo.reload_if_modified()
            self._refresh_jobs_done_counter()
                        

        except Exception as e:
            raise RobotRuntimeFault(
                message=f"_startup_sequence failed: {e}",
                phase="startup",
                error_code="CODE_ERROR",
                handover_file=handover_file,
                cause=e,
            ) from e

    def _refresh_jobs_done_counter(self, job_id=None):
        try:
            count = self.audit.count_done_jobs_today()
            self.ui.post_jobs_done_today(count)
        except Exception as err:
            self.logger.system(err, job_id)

    def _handle_state_change(self, handover_file: HandoverFile) -> None:
        state = handover_file.state
        self._update_dashboard_status(state)

        if state == self.prev_state:
            return

        transition_message=f"state transition detected by CPU-poll: {self.prev_state} -> {state}"

        if not self.handover.is_valid_observed_transition(self.prev_state, state):
            raise RobotRuntimeFault(
                message=f"invalid {transition_message}",
                phase="startup",
                error_code="CODE_ERROR",
                handover_file=handover_file,
            )

        self.logger.system(transition_message, handover_file.job_id)

        if state == "job_running":
            self.audit.mark_running(handover_file.job_id)
        
        self.prev_state = state

    def _map_state_to_phase(self, state: HandoverState) -> RuntimePhase:
        if state == "idle":
            return "poll_intake"

        if state == "job_queued":
            return "waiting_rpa_claim"

        if state == "job_running":
            return "waiting_rpa_execution"

        if state == "job_verifying":
            return "verification"

        if state == "safestop":
            # only RPA tool sets this
            return "waiting_rpa_execution"
           
        raise ValueError(f"unknown state {state}")

    def _enforce_timeouts(self, handover_file):
        state = handover_file.state
        now = time.time()

        if state == "job_queued":
            if self.rpa_tool_claim_started_at is None:
                self.rpa_tool_claim_started_at = now
                self.rpa_tool_execution_started_at = None
                return

            if now - self.rpa_tool_claim_started_at > self.config.rpa_tool_claim_timeout:
                raise RobotRuntimeFault(
                    message="timeout ",
                    phase="waiting_rpa_claim",
                    error_code="PRE_HANDOVER_CRASH",
                    handover_file=handover_file,
                )

        elif state == "job_running":
            if self.rpa_tool_execution_started_at is None:
                self.rpa_tool_claim_started_at = None
                self.rpa_tool_execution_started_at = now
                return
        
            if now - self.rpa_tool_execution_started_at > self.config.rpa_tool_execution_timeout:
                raise RobotRuntimeFault(
                    message="timeout",
                    phase="waiting_rpa_execution",
                    error_code="RPA_TOOL_CRASH",
                    handover_file=handover_file,
                )
        else:
            self.rpa_tool_claim_started_at = None
            self.rpa_tool_execution_started_at = None

    def _update_dashboard_status(self, state=None) -> None:
               
        if state is not None and state not in get_args(HandoverState):
            raise ValueError(f"unknown state: {state}")

        if state == "safestop":
            dashboard_status = "safestop"

        elif state in ("job_queued", "job_running", "job_verifying"):
            dashboard_status = "working"

        elif self.network_service.network_state is False:
            dashboard_status = "no_network"

        elif not self._is_within_operating_schedule():
            dashboard_status = "out_of_office"

        else:
            dashboard_status = "online"

        if dashboard_status == self.dashboard_status:
            return

        self.dashboard_status = dashboard_status
        self.ui.post_status_update(dashboard_status)

    def _run_job_intake(self) -> None:
        ''' job intake logic '''
        try:
            
            # 1. Mail first (priority)
            if self.mail_flow.poll_once():                
                return
            
            # 2. Query (or other scheduled) jobs
            now = time.time()
            if now > self.next_queryflow_check_time:
                if self.query_flow.poll_once():
                    return
                
                self.next_queryflow_check_time = now + self.config.queryflow_poll_interval 
            return

        except RobotRuntimeFault:
            raise
        except Exception as e:
            raise RobotRuntimeFault(
                message=str(e),
                phase="poll_intake",
                error_code="PRE_HANDOVER_CRASH",
                cause=e,
            ) from e

    def _is_within_operating_schedule(self) -> bool:
        now = datetime.datetime.now()

        if now.weekday() not in self.config.operating_days:
            return False

        start = datetime.time(self.config.operating_hours_start) # format datetime.time(hh,mm)
        end = datetime.time(self.config.operating_hours_end) 
        return start <= now.time() <= end

    def _validate_job_handlers_registry(self) -> None:
        # will not catch dublicate names
        for key, handler in self.job_handlers.items():
            self.logger.system(f"loading job handler '{key}', type={str(handler).split()[0][1:]}")
            if not isinstance(key, str) or not key.strip():
                raise ValueError(f"invalid handler key: {key}")

            if not hasattr(handler, "job_name"):
                raise ValueError(f"handler {handler} missing job_name")

            if handler.job_name != key:
                raise ValueError(
                    f"handler registry mismatch: key={key}, handler.job_name={handler.job_name}"
                )   
            
            required_methods = ["precheck_and_build_payload", "verify_result"]
            for method_name in required_methods:
                if not callable(getattr(handler, method_name, None)):
                    raise ValueError(f"handler {key} missing method {method_name}()")

    def _check_for_stop_flag(self):
        ''' to stop main.py on operator manual stop on RPA tool '''

        stopflag = "stop.flag"
 
        if os.path.isfile(stopflag):
            try: os.remove(stopflag)
            except Exception as err: self.logger.system(f"{err}")

            self.logger.system(f"found {stopflag}, initiating shutdown sequence")
            
            try: self.ui.post_shutdown() # request soft exit
            except Exception: os._exit(1)
            
            time.sleep(3)
            os._exit(0)  # kill if still alive after 3 sec 


def main() -> None:
    '''run UI in main thread and the rest async'''
    config = load_or_create_config()

    ui = DashboardUI()
    robot_runtime = RobotRuntime(ui, config)

    threading.Thread(target=robot_runtime.runtime_loop, daemon=True).start()
    ui.run()


if __name__ == "__main__":
    main()